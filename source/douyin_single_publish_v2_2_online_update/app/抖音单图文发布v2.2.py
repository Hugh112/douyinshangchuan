# -*- coding: utf-8 -*-
"""
抖音单图文发布v2.2 - Consolidated
功能：
1. GUI 前端配置浏览器、图片、Excel、定时、等待、上传检测、重试等。
2. 自动打开抖音创作者平台图文发布页。
3. 上传图片后，以"编辑图片 / 已添加1张图片 / 封面设置 / 预览图文"等判断上传完成。
4. 填写文案和话题。
5. 只使用平台音乐面板，悬停音乐后点击"使用"。
6. 音乐成功后直接下滑到发布设置。
7. 支持定时发布 / 立即发布开关。
8. 点击发布前再次校验图片、文案、音乐、发布设置。
9. 支持随机使用图片（运行选项中勾选即可）。
"""

import csv
import hashlib
import json
import os
import queue
import random
import re
import subprocess
import sys
import threading
import time
import urllib.request
import ctypes
from datetime import datetime
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

import pandas as pd
from openpyxl import load_workbook

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
    from tkinter.scrolledtext import ScrolledText
except Exception:
    tk = None

from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError

APP_DIR = Path(__file__).resolve().parent
CONFIG_PATH = APP_DIR / "douyin_gui_config.json"
APP_VERSION = "2.2.1"
APP_NAME = f"抖音单图文发布v{APP_VERSION}"
LOGO_ICO = APP_DIR / "assets" / "app_logo.ico"
LOGO_PNG = APP_DIR / "assets" / "app_logo.png"
ACTIVE_CFG = None

UPDATE_MANIFEST_URL = "https://raw.githubusercontent.com/Hugh112/douyinshangchuan/main/version.json"
UPDATE_TIMEOUT_SECONDS = 8
PRESERVE_UPDATE_PATHS = [
    "app/douyin_gui_config.json",
]


def parse_version(v):
    nums = []
    for part in re.findall(r"\d+", str(v or "0"))[:4]:
        try:
            nums.append(int(part))
        except Exception:
            nums.append(0)
    while len(nums) < 4:
        nums.append(0)
    return tuple(nums)


def fetch_update_manifest():
    req = urllib.request.Request(
        UPDATE_MANIFEST_URL,
        headers={"User-Agent": f"DouyinPublisher/{APP_VERSION}"},
    )
    with urllib.request.urlopen(req, timeout=UPDATE_TIMEOUT_SECONDS) as resp:
        data = resp.read().decode("utf-8-sig", errors="replace")
    manifest = json.loads(data)
    if not isinstance(manifest, dict):
        raise ValueError("版本配置文件格式不正确")
    return manifest


def available_update_info():
    manifest = fetch_update_manifest()
    latest = str(manifest.get("latest_version") or manifest.get("version") or "").strip()
    download_url = str(manifest.get("download_url") or "").strip()
    if not latest:
        raise ValueError("version.json 缺少 latest_version")
    if parse_version(latest) <= parse_version(APP_VERSION):
        return None
    if not download_url:
        raise ValueError("version.json 缺少 download_url")
    return manifest


def format_release_notes(manifest):
    notes = manifest.get("release_notes") or manifest.get("notes") or ""
    if isinstance(notes, list):
        notes = "\n".join("- " + str(x) for x in notes)
    return str(notes).strip()



def default_config():
    user = Path(os.environ.get("USERPROFILE", str(Path.home())))
    one = os.environ.get("OneDrive")
    browser = Path(r"D:\浏览器2.lnk")
    if not browser.exists():
        browser = user / "Desktop" / "浏览器2.lnk"
    if one and not browser.exists():
        browser = Path(one) / "Desktop" / "浏览器2.lnk"
    return {
        "creator_url": "https://creator.douyin.com/",
        "browser_path": str(browser),
        "image_dir": r"D:\抖音\河北",
        "excel_path": r"D:\文案.xlsx",
        "sheet_name": "sheet1",
        "publish_date": "2026-06-21",
        "publish_start_date": "2026-06-21",
        "publish_end_date": "2026-06-21",
        "start_hour": 8,
        "end_hour": 18,
        "per_hour_count": 3,
        "custom_minutes": "5,25,45",
        "wait_min_seconds": 2,
        "wait_max_seconds": 5,
        "publish_interval_min_seconds": 30,
        "publish_interval_max_seconds": 90,
        "upload_check_interval_seconds": 3,
        "upload_max_wait_seconds": 180,
        "topic_wait_seconds": 3,
        "creator_center_wait_seconds": 8,
        "retry_times": 2,
        "cdp_port": 9222,
        "browser_user_data_dir": "",
        "browser_profile_directory": "",
        "reuse_existing_cdp": True,
        "close_chrome_before_start": False,
        "no_raise_browser": True,
        "max_posts_this_run": 0,
        "state_path": r"D:\抖音\河北_publish_state.json",
        "log_path": r"D:\抖音\河北_publish_log.csv",
        "debug_dir": r"D:\抖音\debug_screenshots",
        "delete_image_after_success": True,
        "delete_copy_after_success": True,
        "use_schedule": True,
        "music_required": True,
        "keep_browser_open": True,
        "auto_switch_graphic_mode": True,
        "platform_music_only": True,
        "random_image": False,
    }


def load_config():
    cfg = default_config()
    if CONFIG_PATH.exists():
        try:
            cfg.update(json.loads(CONFIG_PATH.read_text(encoding="utf-8")))
        except Exception:
            pass
    return cfg


def save_config(cfg):
    CONFIG_PATH.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")


def wlog(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)

def popup_pause_alert(msg="脚本已暂停，请查看前端日志。"):
    """
    脚本暂停/自动停止时弹出 Windows 提示框。
    只做提醒，不终止脚本进程。
    """
    try:
        ctypes.windll.user32.MessageBoxW(
            None,
            str(msg),
            APP_NAME,
            0x00000040 | 0x00001000
        )
    except Exception:
        pass


def pause_flag_path(cfg):
    try:
        state = Path(cfg.get("state_path", "douyin_publish_state.json"))
        return state.with_suffix(".pause")
    except Exception:
        return APP_DIR / "douyin_pause.flag"


def wait_if_paused(cfg):
    """
    前端点击"暂停/继续"时使用。
    注意：这里只等待暂停标记，不 terminate 进程，避免中途强断 Playwright 导致 EPIPE。
    """
    try:
        flag = pause_flag_path(cfg)
        shown = False
        while flag.exists():
            if not shown:
                msg = "脚本已暂停。点击前端'暂停/继续'后会继续运行。"
                wlog(msg)
                popup_pause_alert(msg)
                shown = True
            time.sleep(1)
        if shown:
            wlog("暂停已解除，继续运行。")
    except Exception:
        pass


def alert_auto_pause(msg):
    """自动暂停/结束时调用，弹窗提醒一次。"""
    wlog(msg)
    popup_pause_alert(msg)

def step_wait(cfg=None, reason=""):
    cfg = cfg or ACTIVE_CFG or {}
    if cfg:
        wait_if_paused(cfg)
    a = float(cfg.get("wait_min_seconds", 2) or 2)
    b = float(cfg.get("wait_max_seconds", 5) or 5)
    if a < 0:
        a = 0
    if b < a:
        b = a
    d = random.uniform(a, b) if b > a else a
    if reason:
        wlog(f"等待 {d:.1f} 秒：{reason}")
    time.sleep(d)


def worker_python_executable():
    exe = Path(sys.executable)
    if exe.name.lower() == "pythonw.exe":
        py = exe.with_name("python.exe")
        if py.exists():
            return str(py)
    return str(exe)


def no_window_creationflags():
    return getattr(subprocess, "CREATE_NO_WINDOW", 0) if os.name == "nt" else 0


def run_ps(code):
    p = subprocess.run(["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", code],
                       capture_output=True, text=True, encoding="utf-8", errors="ignore")
    if p.returncode != 0:
        raise RuntimeError(p.stderr.strip() or p.stdout.strip())
    return p.stdout.strip()


def resolve_lnk(path):
    ps = rf"""
    $s=(New-Object -COM WScript.Shell).CreateShortcut('{str(path)}');
    [PSCustomObject]@{{TargetPath=$s.TargetPath; Arguments=$s.Arguments}} | ConvertTo-Json -Compress
    """
    data = json.loads(run_ps(ps))
    return data.get("TargetPath", ""), data.get("Arguments", "")


def wait_cdp(port, seconds=25):
    url = f"http://127.0.0.1:{port}/json/version"
    end = time.time() + seconds
    while time.time() < end:
        try:
            with urllib.request.urlopen(url, timeout=1.2) as r:
                if r.status == 200:
                    return True
        except Exception:
            time.sleep(0.6)
    return False


def safe_bring_to_front(cfg, page):
    """
    后台运行模式：默认不主动拉起/置顶浏览器。
    需要看窗口时，把配置 no_raise_browser 改为 False。
    """
    try:
        if cfg and cfg.get("no_raise_browser", True):
            return
        page.bring_to_front()
    except Exception:
        pass


def sanitize_chrome_args(args):
    """删除快捷方式里可能冲突的调试端口参数。"""
    args = str(args or "")
    patterns = [
        r"\s*--remote-debugging-port(?:=|\s+)\"[^\"]*\"",
        r"\s*--remote-debugging-port(?:=|\s+)\S+",
        r"\s*--remote-debugging-address(?:=|\s+)\"[^\"]*\"",
        r"\s*--remote-debugging-address(?:=|\s+)\S+",
    ]
    for pat in patterns:
        args = re.sub(pat, "", args, flags=re.I)
    return args.strip()


def strip_profile_args(args):
    """当用户在前端指定用户目录/Profile时，删除快捷方式里旧的同类参数，避免重复。"""
    args = str(args or "")
    patterns = [
        r"\s*--user-data-dir(?:=|\s+)\"[^\"]*\"",
        r"\s*--user-data-dir(?:=|\s+)\S+",
        r"\s*--profile-directory(?:=|\s+)\"[^\"]*\"",
        r"\s*--profile-directory(?:=|\s+)\S+",
    ]
    for pat in patterns:
        args = re.sub(pat, "", args, flags=re.I)
    return args.strip()


def has_user_data_dir_arg(args):
    return "--user-data-dir" in str(args or "").lower()


def kill_chrome_residue(cfg):
    if not cfg.get("close_chrome_before_start", False):
        return
    wlog("已启用：启动前关闭 Chrome / node 残留进程。")
    for proc in ["chrome.exe", "node.exe"]:
        try:
            subprocess.run(["taskkill", "/F", "/IM", proc], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        except Exception:
            pass

def open_browser_with_cdp(cfg, force_new=False):
    """
    v2.1：稳定 CDP 后台版。
    - 用户不用改现有浏览器快捷方式目标；脚本会在启动时临时追加 --remote-debugging-port。
    - 可在前端填写 browser_user_data_dir / browser_profile_directory，脚本会临时追加，不修改 .lnk 文件本身。
    - 自动清理重复的 --remote-debugging-port，避免快捷方式和脚本重复写端口。
    """
    port = int(cfg.get("cdp_port", 9222))
    kill_chrome_residue(cfg)

    if wait_cdp(port, 2):
        if cfg.get("reuse_existing_cdp", True):
            wlog(f"已检测到浏览器调试端口 {port}，复用现有调试连接。")
            return
        raise RuntimeError(f"调试端口 {port} 已被占用。请关闭旧 Chrome，或勾选'启动前关闭Chrome残留'。")

    browser_path = Path(cfg["browser_path"])
    if not browser_path.exists():
        raise FileNotFoundError(f"浏览器路径不存在：{browser_path}")

    target_url = cfg.get("creator_url", "https://creator.douyin.com/")
    target = str(browser_path)
    args = ""

    if browser_path.suffix.lower() == ".lnk":
        target, args = resolve_lnk(browser_path)
        if not target:
            raise RuntimeError(f"无法解析浏览器快捷方式：{browser_path}")

    args = sanitize_chrome_args(args)

    user_dir = str(cfg.get("browser_user_data_dir", "") or "").strip()
    profile_dir = str(cfg.get("browser_profile_directory", "") or "").strip()

    # 如果前端指定了用户目录/Profile，就以这个为准；不会修改原快捷方式，只是本次启动临时追加参数。
    if user_dir:
        args = strip_profile_args(args)
        args = (args + f' --user-data-dir="{user_dir}"').strip()
        if profile_dir:
            args = (args + f' --profile-directory="{profile_dir}"').strip()
    else:
        # 如果快捷方式本身没有 user-data-dir，为了适配新版 Chrome 远程调试限制，使用脚本专用默认目录。
        if not has_user_data_dir_arg(args):
            default_profile = str(Path(os.environ.get("USERPROFILE", "D:")) / "douyin_chrome_profile_v21")
            args = (args + f' --user-data-dir="{default_profile}"').strip()
            wlog(f"快捷方式未带 user-data-dir，已临时使用脚本专用目录：{default_profile}")

    cmd = f'"{target}" {args} --remote-debugging-port={port} "{target_url}"'
    wlog("启动浏览器命令已生成：不会修改你的浏览器快捷方式目标，只临时追加参数。")
    wlog(f"打开浏览器：{target}")
    subprocess.Popen(cmd, shell=True)
    if not wait_cdp(port, 45):
        raise RuntimeError("浏览器已尝试启动，但无法连接调试端口。请检查用户目录是否被其它 Chrome 占用，或勾选启动前关闭Chrome残留。")



def read_state(cfg):
    p = Path(cfg["state_path"])
    if p.exists():
        try:
            return json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {"copy_index": 0, "slot_index": 0}


def write_state(cfg, state):
    p = Path(cfg["state_path"])
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def append_log(cfg, row):
    p = Path(cfg["log_path"])
    p.parent.mkdir(parents=True, exist_ok=True)
    exists = p.exists()
    with p.open("a", newline="", encoding="utf-8-sig") as f:
        fields = ["time", "image", "copy_index", "schedule_time", "copy_preview", "status"]
        wr = csv.DictWriter(f, fieldnames=fields)
        if not exists:
            wr.writeheader()
        wr.writerow(row)



def schedule_signature(cfg):
    """
    v2.1：排期签名支持多日定时。
    只要日期范围、小时、分钟点、每小时数量或定时开关变化，就重置 slot_index。
    """
    keys = [
        "publish_start_date",
        "publish_end_date",
        "publish_date",
        "start_hour",
        "end_hour",
        "custom_minutes",
        "per_hour_count",
        "use_schedule",
    ]
    return "|".join(f"{k}={str(cfg.get(k, ''))}" for k in keys)

def build_slots(cfg):
    """
    v2.1：支持一次运行多日定时。
    新字段：publish_start_date / publish_end_date。
    兼容旧字段 publish_date：如果没有填写开始/结束日期，就按旧字段单日排期。
    """
    from datetime import timedelta

    start_date_raw = str(cfg.get("publish_start_date") or cfg.get("publish_date") or "").strip()
    end_date_raw = str(cfg.get("publish_end_date") or start_date_raw).strip()
    if not start_date_raw:
        raise RuntimeError("开始日期不能为空，例如 2026-06-21。")
    if not end_date_raw:
        end_date_raw = start_date_raw

    try:
        start_date = datetime.strptime(start_date_raw, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date_raw, "%Y-%m-%d").date()
    except Exception:
        raise RuntimeError("开始日期/结束日期格式必须是 YYYY-MM-DD，例如 2026-06-21。")

    if end_date < start_date:
        raise RuntimeError("结束日期不能早于开始日期。")

    try:
        start = int(cfg.get("start_hour", 8))
        end = int(cfg.get("end_hour", 18))
    except Exception:
        raise RuntimeError("开始小时和结束小时必须填写数字，例如 8、18、23、24。")

    if start < 0 or start > 23:
        raise RuntimeError("开始小时只能填写 0-23，例如 8 或 17。")
    if end < 0 or end > 24:
        raise RuntimeError("结束小时只能填写 0-24。填 24 表示当天 24 点前。")

    end_for_range = 23 if end == 24 else end
    per = max(1, int(cfg.get("per_hour_count", 3)))
    raw = str(cfg.get("custom_minutes", "")).strip()
    if raw:
        mins = []
        for x in re.split(r"[，,\s]+", raw):
            if not x.strip():
                continue
            mi = int(x)
            if mi < 0 or mi > 59:
                raise RuntimeError(f"分钟点只能填写 0-59，当前填写了：{mi}")
            mins.append(mi)
        mins = sorted(set(mins))
    else:
        step = 60 // per
        mins = [min(59, 5 + i * step) for i in range(per)]

    if not mins:
        raise RuntimeError("分钟点不能为空，例如填写：5,25,45")

    slots = []
    day = start_date
    while day <= end_date:
        if end_for_range >= start:
            for h in range(start, end_for_range + 1):
                for mi in mins[:per]:
                    slots.append(datetime(day.year, day.month, day.day, h, mi))
        else:
            # 跨夜排期：例如 20 到 2，会生成当天 20-23 + 次日 0-2。
            for h in range(start, 24):
                for mi in mins[:per]:
                    slots.append(datetime(day.year, day.month, day.day, h, mi))
            next_day = day + timedelta(days=1)
            for h in range(0, end_for_range + 1):
                for mi in mins[:per]:
                    slots.append(datetime(next_day.year, next_day.month, next_day.day, h, mi))
        day = day + timedelta(days=1)

    if not slots:
        raise RuntimeError("没有生成任何定时时间，请检查日期、开始小时、结束小时、每小时条数和分钟点。")
    wlog(f"已生成多日排期 {len(slots)} 个：从 {slots[0].strftime('%Y-%m-%d %H:%M')} 到 {slots[-1].strftime('%Y-%m-%d %H:%M')}")
    return slots


def list_images(cfg):
    d = Path(cfg["image_dir"])
    if not d.exists():
        raise FileNotFoundError(f"图片文件夹不存在：{d}")
    exts = {".jpg", ".jpeg", ".png", ".webp", ".bmp"}
    arr = [p for p in d.iterdir() if p.is_file() and p.suffix.lower() in exts]
    arr.sort(key=lambda p: p.name)
    return arr


def read_copies(cfg):
    excel_value = str(cfg.get("excel_path", "") or "").strip().strip('"')
    if not excel_value:
        raise RuntimeError("文案Excel未设置：请在前端“文案Excel”选择真正的 .xlsx 或 .xls 文案表。")
    p = Path(excel_value)
    if not p.exists():
        raise FileNotFoundError(f"Excel 不存在：{p}")

    suffix = p.suffix.lower()
    allowed = {".xlsx", ".xls", ".xlsm"}
    if suffix not in allowed:
        raise RuntimeError(
            "文案Excel文件格式不正确。\n"
            f"当前选择：{p}\n"
            f"当前后缀：{suffix or '无后缀'}\n"
            "请不要选择软件压缩包、程序目录或其他文件；请在前端“文案Excel”重新选择 .xlsx / .xls / .xlsm 文案表。"
        )

    try:
        xls = pd.ExcelFile(p)
    except Exception as e:
        raise RuntimeError(
            "无法读取文案Excel。\n"
            f"当前选择：{p}\n"
            "请确认该文件是真正的 Excel 表格，且没有损坏；建议另存为 .xlsx 后重新选择。\n"
            f"原始错误：{repr(e)}"
        ) from e

    sheet = next((s for s in xls.sheet_names if s.lower() == str(cfg["sheet_name"]).lower()), None)
    if not sheet:
        raise RuntimeError(f"找不到工作表：{cfg['sheet_name']}；当前工作表：{xls.sheet_names}")
    df = pd.read_excel(p, sheet_name=sheet).dropna(how="all")
    if df.empty:
        return []
    col = next((c for c in df.columns if "文案" in str(c)), df.columns[0])
    return [str(v).strip() for v in df[col].tolist() if not pd.isna(v) and str(v).strip()]


def normalize_copy_text(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def copy_key(value):
    """
    V28：文案唯一标记。用于 Excel 被占用无法删除时，临时跳过已发布文案，避免重复发。
    """
    text = normalize_copy_text(value)
    return hashlib.sha1(text.encode("utf-8", errors="ignore")).hexdigest()


def filter_used_copies(copies, state):
    used = set(state.get("used_copy_keys", []) or [])
    if not used:
        return copies
    filtered = [c for c in copies if copy_key(c) not in used]
    skipped = len(copies) - len(filtered)
    if skipped:
        wlog(f"已根据进度文件跳过 {skipped} 条已发布但未能从 Excel 删除的文案。")
    return filtered


def mark_copy_used_in_state(cfg, copy_text):
    state = read_state(cfg)
    used = list(state.get("used_copy_keys", []) or [])
    key = copy_key(copy_text)
    if key not in used:
        used.append(key)
    state["used_copy_keys"] = used[-5000:]
    write_state(cfg, state)
    wlog("已把当前文案写入进度文件的已用文案列表，避免后续重复使用。")


def delete_copy_from_excel(cfg, copy_text):
    """
    V28：发布成功后从 Excel 中删除已使用文案所在行。
    如果 Excel 被 WPS/Excel 打开占用，会明确提示，并返回 False。
    """
    p = Path(cfg["excel_path"])
    if not p.exists():
        wlog(f"提醒：Excel 不存在，无法删除已用文案：{p}")
        return False

    target = normalize_copy_text(copy_text)
    if not target:
        wlog("提醒：当前文案为空，跳过删除 Excel 文案。")
        return False

    try:
        wb = load_workbook(p)
    except PermissionError:
        wlog(f"删除文案失败：Excel 文件被占用，无法打开写入：{p}。请关闭 WPS/Excel 里的这个文件。")
        return False
    except Exception as e:
        wlog(f"删除文案失败：无法打开 Excel：{repr(e)}")
        return False

    sheet_name = next((s for s in wb.sheetnames if s.lower() == str(cfg["sheet_name"]).lower()), None)
    if not sheet_name:
        wlog(f"提醒：找不到工作表，无法删除已用文案：{cfg.get('sheet_name')}")
        return False

    ws = wb[sheet_name]

    copy_col = None
    for c in range(1, ws.max_column + 1):
        header = normalize_copy_text(ws.cell(row=1, column=c).value)
        if "文案" in header:
            copy_col = c
            break

    if copy_col is None:
        for c in range(1, ws.max_column + 1):
            if any(normalize_copy_text(ws.cell(row=r, column=c).value) for r in range(1, min(ws.max_row, 20) + 1)):
                copy_col = c
                break

    if copy_col is None:
        wlog("提醒：Excel 中没有找到可删除的文案列。")
        return False

    candidate_rows = list(range(2, ws.max_row + 1)) + [1]

    matched_row = None
    for r in candidate_rows:
        value = normalize_copy_text(ws.cell(row=r, column=copy_col).value)
        if value == target:
            matched_row = r
            break

    if matched_row is None:
        target_head = target[:40]
        if len(target_head) >= 10:
            for r in candidate_rows:
                value = normalize_copy_text(ws.cell(row=r, column=copy_col).value)
                if value[:40] == target_head:
                    matched_row = r
                    break

    if matched_row is None:
        wlog("提醒：未在 Excel 中找到匹配的已用文案，未删除。")
        return False

    try:
        ws.delete_rows(matched_row, 1)
        wb.save(p)
        wlog(f"已删除 Excel 中已用文案：第 {matched_row} 行。")
        return True
    except PermissionError:
        wlog(f"删除文案失败：Excel 文件被占用，无法保存：{p}。请关闭 WPS/Excel 里的这个文件。")
        return False
    except Exception as e:
        wlog(f"删除文案失败：保存 Excel 时出错：{repr(e)}")
        return False


def save_debug(cfg, page, name):
    try:
        d = Path(cfg["debug_dir"])
        d.mkdir(parents=True, exist_ok=True)
        path = d / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{name}.png"
        page.screenshot(path=str(path), full_page=True)
        wlog(f"调试截图：{path}")
    except Exception as e:
        wlog(f"截图失败：{e}")


def has_text(page, text, timeout=500):
    try:
        return page.get_by_text(text, exact=False).first.is_visible(timeout=timeout)
    except Exception:
        return False


def has_any(page, texts, timeout=500):
    return any(has_text(page, t, timeout) for t in texts)


def click_text(page, texts, timeout=3500):
    for t in texts:
        try:
            loc = page.get_by_text(t, exact=False).first
            loc.wait_for(state="visible", timeout=timeout)
            loc.click(force=True, timeout=timeout)
            wlog(f"点击：{t}")
            step_wait(reason=f"点击 {t} 后等待")
            return True
        except Exception:
            pass
    for t in texts:
        try:
            loc = page.locator("button").filter(has_text=t).first
            loc.wait_for(state="visible", timeout=1500)
            loc.click(force=True, timeout=timeout)
            wlog(f"点击按钮：{t}")
            step_wait(reason=f"点击按钮 {t} 后等待")
            return True
        except Exception:
            pass
    return False


def first_visible(page, selector):
    try:
        loc = page.locator(selector)
        for i in range(min(loc.count(), 100)):
            item = loc.nth(i)
            try:
                if item.is_visible(timeout=250):
                    return item
            except Exception:
                pass
    except Exception:
        pass
    return None


def visible_count(page, selector):
    try:
        loc = page.locator(selector)
        n = 0
        for i in range(min(loc.count(), 150)):
            try:
                if loc.nth(i).is_visible(timeout=180):
                    n += 1
            except Exception:
                pass
        return n
    except Exception:
        return 0


def wait_login(page):
    if has_any(page, ["扫码登录", "登录", "验证码", "安全验证"], 1200):
        wlog("检测到登录/验证，请在浏览器手动完成。")
        input("完成后按 Enter 继续...")


def image_upload_completed(page):
    """
    V21：严格判断图片是否真正上传完成。
    注意：不能把"封面设置 / 编辑图片 / 选择一张图片作为封面 / 基础信息"当作上传完成，
    因为这些文字在未上传图片时也会出现。

    真正完成标识：
    1. 已添加1张图片 / 已添加 1 张图片 / 已添加N张图片
    2. 清空并重新上传
    3. 编辑图片区块里出现已添加图片数量
    """
    try:
        body = page.locator("body").inner_text(timeout=2500)
    except Exception:
        body = ""

    if any(w in body for w in ["上传失败", "图片处理失败", "重新上传失败"]):
        return False

    # 明确的上传完成文字
    if re.search(r"已添加\s*\d+\s*张图片", body):
        return True

    if "清空并重新上传" in body:
        return True

    # DOM 层面：只在编辑图片/封面相关区域里确认"已添加N张图片"
    try:
        ok = page.evaluate("""
        () => {
          function visible(el){
            const r=el.getBoundingClientRect();
            const s=getComputedStyle(el);
            return r.width>0 && r.height>0 && s.display!=='none' && s.visibility!=='hidden';
          }
          function txt(el){ return ((el.innerText||el.textContent||'')+'').trim(); }

          const bodyText = txt(document.body);
          if (bodyText.includes('上传失败') || bodyText.includes('图片处理失败')) return false;

          const nodes=[...document.querySelectorAll('div,section,span,p,button')].filter(visible);
          for (const el of nodes) {
            const t = txt(el);
            if (!t || t.length > 300) continue;
            if (/已添加\\s*\\d+\\s*张图片/.test(t)) return true;
            if (t.includes('清空并重新上传')) return true;
          }

          return false;
        }
        """)
        return bool(ok)
    except Exception:
        return False


def wait_uploaded_with_config(cfg, page):
    interval = float(cfg.get("upload_check_interval_seconds", 3) or 3)
    max_wait = float(cfg.get("upload_max_wait_seconds", 180) or 180)
    interval = max(1, interval)
    max_wait = max(interval, max_wait)
    wlog(f"判断图片是否上传完成，检测间隔={interval}秒，最大等待={max_wait}秒。")
    start = time.time()
    while True:
        if image_upload_completed(page):
            wlog("图片上传完成：检测到编辑图片/已添加图片/封面设置等完成标识。")
            return True
        if has_any(page, ["上传失败", "图片处理失败", "重新上传"], 800):
            raise RuntimeError("检测到图片上传失败。")
        if time.time() - start >= max_wait:
            raise RuntimeError(f"等待图片上传超过 {max_wait} 秒，仍未检测到完成标识。")
        wlog(f"暂未检测到图片上传完成，等待 {interval:.1f} 秒后重试。")
        time.sleep(interval)


def validate_uploaded_image(cfg, page):
    ok = image_upload_completed(page)
    if not ok:
        save_debug(cfg, page, "validate_uploaded_image_failed")
        raise RuntimeError("图片上传状态未完成，未检测到编辑图片/已添加图片/封面设置。")
    wlog("校验通过：图片已上传完成。")


def split_text_topics(text):
    """
    从文案里提取话题，并从正文中彻底清除话题残留。
    修复点：
    1. 支持 #话题、##话题、###话题、＃话题 等写法。
    2. 删除正文里因为双 # / 多 # 导致残留的 ###。
    3. 返回去重后的干净话题，不带 #。
    """
    text = str(text or "").replace("＃", "#")

    # 支持一个或多个 #，也支持 # 后面有空格的情况
    topics = re.findall(r"#+\s*([\u4e00-\u9fa5A-Za-z0-9_]+)", text)
    seen = []
    for t in topics:
        t = str(t).strip().lstrip("#").strip()
        if t and t not in seen:
            seen.append(t)

    # 删除完整话题，避免正文里留下单独的 #
    body = re.sub(r"#+\s*[\u4e00-\u9fa5A-Za-z0-9_]+", "", text)
    # 兜底删除所有孤立/残留 #，解决页面出现 ### 的问题
    body = re.sub(r"[#＃]+", "", body)
    body = re.sub(r"[ \t]{2,}", " ", body)
    body = re.sub(r"\n{3,}", "\n\n", body).strip()
    return body, seen


def upload_tab_ready(page):
    """
    V24：判断是否已经在"发布图文"上传页。
    典型页面：creator.douyin.com/creator-micro/content/upload?default-tab=3
    顶部有"发布视频 / 发布图文 / 发布全景视频 / 发布文章"，中间有红色"上传图文"。
    """
    try:
        url = page.url.lower()
    except Exception:
        url = ""
    try:
        text = page.locator("body").inner_text(timeout=2500)
    except Exception:
        text = ""
    if "content/upload" in url and ("发布图文" in text or "上传图文" in text):
        return True
    if "上传图文" in text and "图片格式" in text and "图片大小" in text:
        return True
    if "点击上传" in text and "直接将图片文件拖入此区域" in text:
        return True
    return False


def click_upload_graphic_button_or_input(cfg, page, image_path):
    """
    V24：在发布图文上传页，直接上传图片。
    不处理"继续编辑 / 放弃"，也不重复点击发布入口。
    """
    wlog("当前为发布图文上传页，直接点击上传图文/文件 input 上传图片。")
    try:
        inp = page.locator("input[type='file']")
        chosen = None
        for i in range(min(inp.count(), 30)):
            item = inp.nth(i)
            try:
                accept = (item.get_attribute("accept") or "").lower()
                if "image" in accept or ".jpg" in accept or ".png" in accept or ".jpeg" in accept or ".webp" in accept:
                    chosen = item
                    break
            except Exception:
                pass
        if chosen is None and inp.count() > 0:
            chosen = inp.first
        if chosen is not None:
            chosen.set_input_files(str(image_path), timeout=30000)
            wlog("已通过文件 input 上传图片。")
            return True
    except Exception as e:
        wlog(f"直接设置上传 input 失败，改点上传图文按钮：{repr(e)}")

    upload_words = ["上传图文", "点击上传", "上传图片", "选择图片", "选择文件", "上传"]
    for word in upload_words:
        try:
            with page.expect_file_chooser(timeout=8000) as fc:
                ok = page.evaluate("""
                (word) => {
                  function visible(el){
                    const r = el.getBoundingClientRect();
                    const s = getComputedStyle(el);
                    return r.width > 0 && r.height > 0 && s.display !== 'none' && s.visibility !== 'hidden';
                  }
                  function txt(el){ return ((el.innerText || el.textContent || '') + '').trim(); }
                  const nodes = [...document.querySelectorAll('button,[role=button],span,div,a')].filter(visible);
                  const arr = [];
                  for (const el of nodes) {
                    const t = txt(el);
                    if (!t || t.length > 80) continue;
                    if (!t.includes(word)) continue;
                    const r = el.getBoundingClientRect();
                    let score = 0;
                    if (t === word) score += 50;
                    if (word === '上传图文') score += 100;
                    if (r.top > 150 && r.top < window.innerHeight * 0.8) score += 20;
                    arr.push({el, score});
                  }
                  arr.sort((a,b)=>b.score-a.score);
                  if (arr.length) {
                    (arr[0].el.closest('button') || arr[0].el.closest('[role=button]') || arr[0].el).click();
                    return true;
                  }
                  return false;
                }
                """, word)
                if not ok:
                    raise RuntimeError(f"没找到按钮：{word}")
            fc.value.set_files(str(image_path))
            wlog(f"已点击'{word}'并选择图片。")
            return True
        except Exception:
            pass
    return False

def platform_image_publish_page_ready(page):
    """
    V24：图文发布准备状态包括：
    1. 图文编辑页 post/image
    2. 图文上传页 content/upload?default-tab=3
    """
    if upload_tab_ready(page):
        return True
    try:
        url = page.url.lower()
    except Exception:
        url = ""
    if "post/image" in url or "media_type=image" in url:
        return True
    try:
        body = page.locator("body").inner_text(timeout=2000)
    except Exception:
        body = ""
    return ("作品描述" in body and "封面设置" in body and "选择音乐" in body)


def click_creator_publish_entry(cfg, page):
    """
    V29：不点击会触发新标签的发布入口。
    所有页面跳转都在当前标签里用 page.goto 完成。
    """
    wait_sec = float(cfg.get("creator_center_wait_seconds", 8) or 8)
    if wait_sec < 0:
        wait_sec = 0

    wlog("当前标签打开抖音创作者中心首页。")
    page.goto(cfg.get("creator_url", "https://creator.douyin.com/"), wait_until="domcontentloaded", timeout=60000)
    safe_bring_to_front(cfg, page)

    wait_login(page)

    wlog(f"进入创作者中心后等待 {wait_sec:.1f} 秒，再在当前标签进入发布图文页。")
    time.sleep(wait_sec)

    upload_url = "https://creator.douyin.com/creator-micro/content/upload?default-tab=3"
    page.goto(upload_url, wait_until="domcontentloaded", timeout=60000)
    step_wait(cfg, "当前标签打开发布图文上传页后等待")

    if platform_image_publish_page_ready(page):
        wlog("已在当前标签进入发布图文页面。")
        return True

    save_debug(cfg, page, "image_publish_page_not_ready")
    raise RuntimeError("没有成功进入图文发布页面。")



def enter_publish_page(cfg, page):
    """
    V22：每次都从创作者中心首页重新进入发布图文页。
    不再保留浏览器原网页直接继续。
    """
    click_creator_publish_entry(cfg, page)
    wait_login(page)


def upload_image(cfg, page, image_path):
    """
    V24：如果当前已经是发布图文上传页，直接点击"上传图文"上传。
    不管"继续编辑 / 放弃"，不重复点击发布入口。
    """
    wlog(f"上传图片：{Path(image_path).name}")
    if upload_tab_ready(page):
        ok = click_upload_graphic_button_or_input(cfg, page, image_path)
        if ok:
            return
        save_debug(cfg, page, "upload_graphic_button_failed")
        raise RuntimeError("当前是发布图文上传页，但没能点击上传图文按钮或上传 input。")

    try:
        inp = page.locator("input[type='file']")
        chosen = None
        for i in range(min(inp.count(), 30)):
            item = inp.nth(i)
            try:
                accept = (item.get_attribute("accept") or "").lower()
                if "image" in accept or ".jpg" in accept or ".png" in accept or ".jpeg" in accept or ".webp" in accept:
                    chosen = item
                    break
            except Exception:
                pass
        if chosen is None and inp.count() > 0:
            chosen = inp.first
        if chosen is not None:
            chosen.set_input_files(str(image_path), timeout=30000)
            wlog("已选择图片。")
            return
    except Exception as e:
        wlog(f"直接设置图片 input 失败，改用文件选择器：{repr(e)}")

    for t in ["上传图文", "点击上传", "上传图片", "选择图片", "添加图片", "选择文件", "上传"]:
        try:
            with page.expect_file_chooser(timeout=8000) as fc:
                click_text(page, [t], timeout=5000)
            fc.value.set_files(str(image_path))
            wlog("已通过文件选择器选择图片。")
            return
        except Exception:
            pass
    save_debug(cfg, page, "upload_entry_not_found")
    raise RuntimeError("没有找到图片上传入口。")


def wait_after_upload_to_editor(cfg, page):
    """
    V24：点击上传图文/选择图片后，等待页面进入图文编辑页。
    """
    max_wait = float(cfg.get("upload_max_wait_seconds", 180) or 180)
    interval = float(cfg.get("upload_check_interval_seconds", 3) or 3)
    interval = max(1, interval)
    start_time = time.time()
    while True:
        try:
            text = page.locator("body").inner_text(timeout=2500)
        except Exception:
            text = ""
        if ("作品描述" in text or "基础信息" in text) and ("选择音乐" in text or "封面设置" in text):
            wlog("已进入图文编辑页。")
            return True
        if re.search(r"已添加\s*\d+\s*张图片", text):
            wlog("检测到已添加图片，继续。")
            return True
        if "上传失败" in text or "图片处理失败" in text:
            raise RuntimeError("图片上传失败。")
        if time.time() - start_time > max_wait:
            save_debug(cfg, page, "wait_editor_after_upload_timeout")
            raise RuntimeError(f"上传后等待进入编辑页超过 {max_wait} 秒。")
        wlog(f"等待进入图文编辑页，{interval:.1f} 秒后重试。")
        time.sleep(interval)

def find_editor(page):
    selectors = ["div[contenteditable='true']", "[contenteditable='true']", "textarea", "[role='textbox']"]
    for sel in selectors:
        item = first_visible(page, sel)
        if item:
            return item
    return None


def fill_copy(cfg, page, text):
    """
    填写正文并点亮话题（逐字输入版，不占用剪贴板）。
    修复点：
    1. 话题每输入一次后等待识别，避免太快导致话题没点亮。
    2. 函数缩进修正，可直接运行。
    3. 使用 page.keyboard.type() 逐字输入，不占用剪贴板。
    """
    body, topics = split_text_topics(text)
    editor = find_editor(page)
    if not editor:
        save_debug(cfg, page, "editor_not_found")
        raise RuntimeError("没有找到文案输入框。")

    editor.click(force=True)
    step_wait(cfg, "点击文案框后等待")

    # 先清空文案框
    page.keyboard.press("Control+A")
    page.keyboard.press("Backspace")
    step_wait(cfg, "清空文案框后等待")

    # 逐字输入正文（不占用剪贴板）
    input_text = body if body else text
    wlog(f"逐字输入正文，共 {len(input_text)} 字")
    page.keyboard.type(input_text, delay=0)
    step_wait(cfg, "输入文案后等待")

    if topics:
        wlog("点亮话题：" + "、".join(topics))
        page.keyboard.press("End")
        page.keyboard.press("Enter")

        # 等待抖音话题输入区域稳定
        time.sleep(float(cfg.get("topic_first_wait_seconds", 1.5) or 1.5))

        # 每个话题单独输入，输入后等待平台识别，再按 Enter 点亮
        topic_wait = float(cfg.get("topic_wait_seconds", 2.5) or 2.5)
        topic_after_enter_wait = float(cfg.get("topic_after_enter_wait_seconds", 1) or 1)

        for topic in topics:
            topic = str(topic).strip().replace("＃", "#").lstrip("#").strip()
            if not topic:
                continue
            # 逐字输入话题 #tag（不占用剪贴板）
            page.keyboard.type("#" + topic, delay=0)

            wlog(f"等待识别话题：#{topic}")
            time.sleep(topic_wait)

            page.keyboard.press("Enter")
            time.sleep(topic_after_enter_wait)
            page.keyboard.type(" ")


def validate_copy_filled(cfg, page, text):
    body, _ = split_text_topics(text)
    key = (body or text)[:18].strip()
    try:
        page_text = page.locator("body").inner_text(timeout=3000)
    except Exception:
        page_text = ""
    if key and key not in page_text:
        save_debug(cfg, page, "validate_copy_filled_failed")
        raise RuntimeError("文案未检测到，可能未填写成功。")
    wlog("校验通过：文案已写入。")


def verify_music_selected(page):
    """
    V20：严格判断音乐是否真的使用成功。
    只看"选择音乐"这一行/扩展信息区域，不能用页面其他地方的 00:xx 时长误判。

    成功条件：
    1. 选择音乐区域出现"修改音乐 / 更换音乐 / 删除音乐"
    2. 或选择音乐区域里没有"点击添加合适作品风格音乐"，且有歌曲名 + 时长

    失败条件：
    选择音乐区域仍然出现"点击添加合适作品风格音乐"
    """
    try:
        result = page.evaluate("""
        () => {
          function visible(el){
            const r = el.getBoundingClientRect();
            const s = getComputedStyle(el);
            return r.width > 0 && r.height > 0 && s.display !== 'none' && s.visibility !== 'hidden';
          }
          function txt(el){ return ((el.innerText || el.textContent || '') + '').trim(); }

          const nodes = [...document.querySelectorAll('div,section,li')].filter(visible);
          const rows = [];

          for (const el of nodes) {
            const t = txt(el);
            if (!t || t.length > 600) continue;
            // 只抓"选择音乐"相关的那一块，避免抓到右侧预览或整页文本
            if (t.includes('选择音乐') || t.includes('点击添加合适作品风格音乐') || t.includes('修改音乐') || t.includes('更换音乐')) {
              const r = el.getBoundingClientRect();
              // 只考虑页面左侧/中间编辑区，不考虑右侧预览手机
              if (r.left > window.innerWidth * 0.75) continue;
              rows.push({
                text: t,
                x: r.left, y: r.top, w: r.width, h: r.height,
                area: r.width * r.height
              });
            }
          }

          // 优先选择面积适中的行块，避免选中整页大容器
          rows.sort((a,b) => {
            const as = (a.text.includes('选择音乐') ? 1000 : 0) - Math.abs(a.area - 60000);
            const bs = (b.text.includes('选择音乐') ? 1000 : 0) - Math.abs(b.area - 60000);
            return bs - as;
          });

          const musicBlock = rows[0] || null;
          const text = musicBlock ? musicBlock.text : '';

          const hasAddPrompt = text.includes('点击添加合适作品风格音乐');
          const hasModify = text.includes('修改音乐') || text.includes('更换音乐') || text.includes('删除音乐');
          const hasDuration = /\\d{1,2}:\\d{2}/.test(text);

          return {
            ok: !hasAddPrompt && (hasModify || hasDuration),
            hasAddPrompt,
            hasModify,
            hasDuration,
            text: text.slice(0, 300),
            rowsCount: rows.length
          };
        }
        """)
    except Exception as e:
        wlog(f"音乐校验异常：{repr(e)}")
        return False

    wlog("音乐校验详情：" + json.dumps(result, ensure_ascii=False)[:600])

    if result.get("hasAddPrompt"):
        wlog("音乐校验失败：选择音乐区域仍显示'点击添加合适作品风格音乐'。")
        return False

    if result.get("ok"):
        wlog("音乐校验通过：选择音乐区域已显示已使用音乐。")
        return True

    return False


def validate_music_added(cfg, page):
    if not verify_music_selected(page):
        save_debug(cfg, page, "validate_music_failed")
        raise RuntimeError("音乐未使用成功：选择音乐区域仍未显示已使用音乐。")
    wlog("校验通过：音乐已选择。")


def platform_music_panel_opened(page):
    """
    严格判断右侧平台音乐面板是否打开。
    不能因为页面本身有"选择音乐"字段就误判。
    """
    try:
        return bool(page.evaluate("""
        () => {
          function visible(el){
            const r = el.getBoundingClientRect();
            const s = getComputedStyle(el);
            return r.width > 0 && r.height > 0 && s.display !== 'none' && s.visibility !== 'hidden';
          }
          function txt(el){ return ((el.innerText || el.textContent || '') + '').trim(); }
          const vw = window.innerWidth;
          const nodes = [...document.querySelectorAll('div,section,[role=dialog]')].filter(visible);
          for (const el of nodes) {
            const r = el.getBoundingClientRect();
            const t = txt(el);
            // 音乐面板在右侧抽屉区
            if (r.left < vw * 0.45) continue;
            if (!t.includes('选择音乐')) continue;
            if (t.includes('热门榜') || t.includes('热门音乐') || t.includes('搜索音乐') || t.includes('原创榜') || t.includes('使用')) {
              return true;
            }
          }
          return false;
        }
        """))
    except Exception:
        return False


def scroll_to_music_area(cfg, page):
    """
    滚动到扩展信息里的选择音乐区域。
    """
    for i in range(1, 6):
        try:
            page.evaluate("""
            () => {
              function visible(el){
                const r = el.getBoundingClientRect();
                const s = getComputedStyle(el);
                return r.width > 0 && r.height > 0 && s.display !== 'none' && s.visibility !== 'hidden';
              }
              function txt(el){ return ((el.innerText || el.textContent || '') + '').trim(); }
              const nodes = [...document.querySelectorAll('div,section,span,label,p,button')].filter(visible);
              const el = nodes.find(n => {
                const t = txt(n);
                return t && t.length < 200 && (t.includes('点击添加合适作品风格音乐') || t.includes('选择音乐') || t.includes('扩展信息'));
              });
              if (el) el.scrollIntoView({block:'center', inline:'nearest'});
            }
            """)
        except Exception:
            pass
        step_wait(cfg, f"滚动到选择音乐区域 第{i}次")
        try:
            text = page.locator("body").inner_text(timeout=2000)
        except Exception:
            text = ""
        if "点击添加合适作品风格音乐" in text or "选择音乐" in text:
            return True
    return False


def open_music_panel(cfg, page):
    if verify_music_selected(page):
        return True
    if platform_music_panel_opened(page):
        wlog("平台音乐面板已经打开。")
        return True

    scroll_to_music_area(cfg, page)

    # 只点击"选择音乐"区域右侧按钮或该行右侧区域
    js_click_music = """
    () => {
      function visible(el){
        const r = el.getBoundingClientRect();
        const s = getComputedStyle(el);
        return r.width > 0 && r.height > 0 && s.display !== 'none' && s.visibility !== 'hidden';
      }
      function txt(el){ return ((el.innerText || el.textContent || '') + '').trim(); }

      const vw = window.innerWidth;
      const nodes = [...document.querySelectorAll('div,section,li')].filter(visible);
      const candidates = [];

      for (const el of nodes) {
        const t = txt(el);
        if (!t) continue;
        if (t.length > 800) continue;
        if (!(t.includes('选择音乐') || t.includes('点击添加合适作品风格音乐'))) continue;
        if (t.includes('热门榜') || t.includes('搜索音乐')) continue; // 排除右侧面板
        const r = el.getBoundingClientRect();
        if (r.left > vw * 0.75) continue; // 排除右侧预览/抽屉
        if (r.height < 30 || r.width < 250) continue;

        let score = 0;
        if (t.includes('点击添加合适作品风格音乐')) score += 100;
        if (t.includes('选择音乐')) score += 60;
        if (t.includes('扩展信息')) score += 10;
        // 面积适中更像音乐行
        score -= Math.abs((r.width * r.height) - 50000) / 10000;

        candidates.push({el, score, text:t.slice(0,120), x:r.left, y:r.top, w:r.width, h:r.height});
      }

      candidates.sort((a,b)=>b.score-a.score);
      const row = candidates[0];
      if (!row) return {ok:false, reason:'music row not found'};

      // 优先点击行内可见的"选择音乐"按钮/文字
      const clickables = [...row.el.querySelectorAll('button,[role=button],span,div,a')].filter(visible);
      const btns = [];
      for (const b of clickables) {
        const t = txt(b);
        const r = b.getBoundingClientRect();
        if (t === '选择音乐' || t.includes('选择音乐')) {
          btns.push({el:b, x:r.left, y:r.top, w:r.width, h:r.height, text:t});
        }
      }
      btns.sort((a,b)=>b.x-a.x);
      if (btns.length) {
        (btns[0].el.closest('button') || btns[0].el.closest('[role=button]') || btns[0].el).click();
        return {ok:true, method:'button', text:btns[0].text};
      }

      // 兜底：点击该行右侧区域，通常是"选择音乐"按钮所在位置
      const x = row.x + row.w - 70;
      const y = row.y + row.h / 2;
      const target = document.elementFromPoint(x, y);
      if (target) {
        target.click();
        return {ok:true, method:'right-area', x, y, rowText:row.text};
      }

      return {ok:false, reason:'no target at right area', rowText:row.text};
    }
    """

    for attempt in range(1, 8):
        try:
            result = page.evaluate(js_click_music)
            wlog("点击音乐入口结果：" + json.dumps(result, ensure_ascii=False)[:400])
            if result and result.get("ok"):
                step_wait(cfg, f"第 {attempt} 次点击选择音乐后等待面板")
                if platform_music_panel_opened(page):
                    wlog("平台音乐面板已打开。")
                    return True
        except Exception as e:
            wlog(f"第 {attempt} 次点击选择音乐入口失败：{repr(e)}")

        # 有些情况下点了行但没打开，再点"选择音乐"文字坐标
        try:
            rects = get_text_rects(page, "选择音乐")
            # 选择左侧/中间编辑区的文字
            rects = [r for r in rects if r.get("x", 9999) < 700]
            if rects:
                r = rects[-1]
                page.mouse.click(float(r["cx"]), float(r["cy"]))
                step_wait(cfg, "点击选择音乐文字后等待")
                if platform_music_panel_opened(page):
                    wlog("平台音乐面板已打开。")
                    return True
        except Exception:
            pass

        scroll_to_music_area(cfg, page)
        step_wait(cfg, "音乐面板未打开，准备重试")

    save_debug(cfg, page, "music_panel_not_opened")
    raise RuntimeError("没有打开平台音乐面板。")


def choose_music(cfg, page):
    """
    V28：强制使用平台音乐面板里的"热门榜"。
    只在右侧音乐抽屉中点击热门榜，并只从右侧音乐列表里悬停歌曲、点击"使用"。
    """
    open_music_panel(cfg, page)
    if verify_music_selected(page):
        return True

    # 只点击右侧抽屉里的"热门榜"
    for attempt in range(1, 5):
        try:
            res = page.evaluate("""
            () => {
              function visible(el){
                const r = el.getBoundingClientRect();
                const s = getComputedStyle(el);
                return r.width > 0 && r.height > 0 && s.display !== 'none' && s.visibility !== 'hidden';
              }
              function txt(el){ return ((el.innerText || el.textContent || '') + '').trim(); }
              const vw = window.innerWidth;
              const nodes = [...document.querySelectorAll('button,[role=button],span,div,a')].filter(visible);
              const arr = [];
              for (const el of nodes) {
                const t = txt(el);
                if (t !== '热门榜' && !t.includes('热门榜')) continue;
                const r = el.getBoundingClientRect();
                if (r.left < vw * 0.45) continue;  // 只要右侧音乐面板
                if (r.top > 220) continue;         // tab 一般在上方
                let score = 0;
                if (t === '热门榜') score += 100;
                if (r.left > vw * 0.55) score += 20;
                arr.push({el, score, text:t, x:r.left, y:r.top});
              }
              arr.sort((a,b)=>b.score-a.score);
              if (arr.length) {
                (arr[0].el.closest('button') || arr[0].el.closest('[role=button]') || arr[0].el).click();
                return {ok:true, text:arr[0].text};
              }
              return {ok:false};
            }
            """)
            wlog("点击热门榜结果：" + json.dumps(res, ensure_ascii=False)[:300])
            step_wait(cfg, "点击热门榜后等待")
            if res and res.get("ok"):
                break
        except Exception as e:
            wlog(f"第 {attempt} 次点击热门榜失败：{repr(e)}")
        step_wait(cfg, "等待热门榜标签")

    # 从右侧热门榜列表选择音乐
    for attempt in range(1, 9):
        rows = page.evaluate("""
        () => {
          function visible(el){
            const r = el.getBoundingClientRect();
            const s = getComputedStyle(el);
            return r.width > 0 && r.height > 0 && s.display !== 'none' && s.visibility !== 'hidden';
          }
          function txt(el){ return ((el.innerText || el.textContent || '') + '').trim(); }
          const vw = window.innerWidth, vh = window.innerHeight;
          const out = [];

          // 找右侧音乐面板范围
          const panels = [...document.querySelectorAll('div,section,[role=dialog]')].filter(visible).map(el => {
            const r = el.getBoundingClientRect();
            const t = txt(el);
            return {el, r, t};
          }).filter(x => x.r.left > vw * 0.45 && x.t.includes('选择音乐') && x.t.includes('热门榜'));

          const panel = panels.sort((a,b)=>(b.r.width*b.r.height)-(a.r.width*a.r.height))[0];
          const candidates = panel ? [...panel.el.querySelectorAll('div,li,[role=listitem]')].filter(visible) : [...document.querySelectorAll('div,li,[role=listitem]')].filter(visible);

          for (const el of candidates) {
            const r = el.getBoundingClientRect();
            const t = txt(el);
            if (r.left < vw * 0.45 || r.top < 90 || r.top > vh - 60 || r.width < 250) continue;
            if (r.height < 35 || r.height > 130) continue;
            if (!t || t.length < 5 || t.length > 260) continue;
            if (t.includes('上传') || t.includes('本地') || t.includes('文件') || t.includes('选择音乐')) continue;
            if (t.includes('推荐') && !t.includes('万')) continue;
            // 热门榜行通常有使用量"万使用"和时长
            if (t.includes('万') || /\\d{1,2}:\\d{2}/.test(t)) {
              out.push({
                cx: r.left + r.width * 0.45,
                cy: r.top + r.height / 2,
                text: t.slice(0, 80),
                top: r.top
              });
            }
          }
          return out.slice(0, 15);
        }
        """)

        if not rows:
            wlog("热门榜音乐列表暂未加载，滚动后重试。")
            try:
                page.mouse.wheel(0, 500)
            except Exception:
                pass
            step_wait(cfg, "等待热门榜音乐列表")
            continue

        random.shuffle(rows)
        for row in rows[:6]:
            wlog("悬停热门榜音乐：" + row.get("text", ""))
            page.mouse.move(float(row["cx"]), float(row["cy"]))
            step_wait(cfg, "悬停热门榜音乐后等待使用按钮")

            btn = page.evaluate("""
            () => {
              function visible(el){
                const r = el.getBoundingClientRect();
                const s = getComputedStyle(el);
                return r.width > 0 && r.height > 0 && s.display !== 'none' && s.visibility !== 'hidden';
              }
              function txt(el){ return ((el.innerText || el.textContent || '') + '').trim(); }
              const vw = window.innerWidth;
              const arr = [];
              for (const el of [...document.querySelectorAll('button,[role=button],span,div')].filter(visible)) {
                const t = txt(el);
                const r = el.getBoundingClientRect();
                if ((t === '使用' || t.startsWith('使用')) && r.left > vw * 0.62 && r.width >= 25 && r.height >= 20) {
                  arr.push({x:r.left+r.width/2, y:r.top+r.height/2, left:r.left});
                }
              }
              arr.sort((a,b)=>b.left-a.left);
              return arr[0] || null;
            }
            """)

            if btn:
                page.mouse.click(float(btn["x"]), float(btn["y"]))
                step_wait(cfg, "点击使用热门榜音乐后等待")
                click_text(page, ["确认使用", "确定", "确认", "完成"], timeout=2500)
                step_wait(cfg, "等待热门榜音乐使用生效")
                if verify_music_selected(page):
                    wlog("热门榜音乐已使用成功。")
                    return True

        try:
            page.mouse.wheel(0, 700)
        except Exception:
            pass
        step_wait(cfg, "滚动热门榜音乐列表后等待")

    save_debug(cfg, page, "hot_music_use_failed")
    raise RuntimeError("热门榜音乐面板已打开，但未能点击使用音乐。")


def close_music_panel_if_open(cfg, page):
    if verify_music_selected(page):
        wlog("音乐已使用成功，直接继续。")
        return
    try:
        page.keyboard.press("Escape")
        step_wait(cfg, "关闭音乐面板后等待")
    except Exception:
        pass


def select_music(cfg, page):
    if verify_music_selected(page):
        return True
    choose_music(cfg, page)
    validate_music_added(cfg, page)
    return True


def validate_publish_settings_visible(cfg, page, raise_error=True):
    """
    V20：校验"发布设置"区域确实在当前可视区域，而不是仅存在于 DOM 文本里。
    必须看到：
    - 发布设置
    - 发布时间/立即发布/定时发布
    - 底部发布按钮区域
    """
    try:
        info = page.evaluate("""
        () => {
          function visible(el){
            const r = el.getBoundingClientRect();
            const s = getComputedStyle(el);
            return r.width > 0 && r.height > 0 && s.display !== 'none' && s.visibility !== 'hidden';
          }
          function txt(el){ return ((el.innerText || el.textContent || '') + '').trim(); }
          function inViewport(r){
            return r.top >= 0 && r.top < window.innerHeight && r.bottom > 0 && r.left < window.innerWidth && r.right > 0;
          }

          let publishSettingVisible = false;
          let publishTimeVisible = false;
          let publishButtonVisible = false;

          const nodes = [...document.querySelectorAll('div,section,span,label,p,button')].filter(visible);
          for (const el of nodes) {
            const t = txt(el);
            if (!t || t.length > 120) continue;
            const r = el.getBoundingClientRect();
            const iv = inViewport(r);

            if (iv && t.includes('发布设置')) publishSettingVisible = true;
            if (iv && (t.includes('发布时间') || t.includes('立即发布') || t.includes('定时发布'))) publishTimeVisible = true;
            if (iv && (t === '发布' || t.includes('发布'))) {
              // 只接受页面底部区域的发布按钮，避免左侧导航"发布"误判
              if (r.top > window.innerHeight * 0.55 && r.left > window.innerWidth * 0.1 && r.left < window.innerWidth * 0.75) {
                publishButtonVisible = true;
              }
            }
          }

          return {
            publishSettingVisible,
            publishTimeVisible,
            publishButtonVisible,
            scrollY: window.scrollY,
            innerHeight: window.innerHeight,
            bodyHeight: document.documentElement.scrollHeight
          };
        }
        """)
    except Exception as e:
        info = {"error": repr(e), "publishSettingVisible": False, "publishTimeVisible": False, "publishButtonVisible": False}

    ok = bool(info.get("publishSettingVisible") and info.get("publishTimeVisible") and info.get("publishButtonVisible"))
    wlog("发布设置可视校验：" + json.dumps(info, ensure_ascii=False)[:500])

    if not ok and raise_error:
        save_debug(cfg, page, "publish_settings_not_visible")
        raise RuntimeError("没有在当前可视区域检测到完整发布设置和底部发布按钮。")
    return ok


def scroll_to_publish_settings(cfg, page):
    """
    V20：强制滚动发布页到最底部。
    抖音页面常见内部容器滚动，不是 window 滚动，所以同时滚动所有可滚动容器。
    只有发布设置和底部发布按钮都进入可视区后，才算成功。
    """
    wlog("强制下滑到发布页最底部/发布设置区域。")

    scroll_js = """
    () => {
      function visible(el){
        const r = el.getBoundingClientRect();
        const s = getComputedStyle(el);
        return r.width > 0 && r.height > 0 && s.display !== 'none' && s.visibility !== 'hidden';
      }
      function txt(el){ return ((el.innerText || el.textContent || '') + '').trim(); }

      // 先找发布设置或发布按钮，尽量滚入视野
      const keys = ['发布设置', '发布时间', '立即发布', '定时发布'];
      const nodes = [...document.querySelectorAll('div,section,span,label,p,button')].filter(visible);
      for (const key of keys) {
        const el = nodes.find(n => {
          const t = txt(n);
          return t && t.length < 120 && t.includes(key);
        });
        if (el) {
          el.scrollIntoView({block:'center', inline:'nearest'});
          break;
        }
      }

      // window 到底
      window.scrollTo(0, document.documentElement.scrollHeight);

      // 所有内部滚动容器到底
      const all = [...document.querySelectorAll('*')];
      for (const el of all) {
        try {
          const r = el.getBoundingClientRect();
          const st = getComputedStyle(el);
          const canScroll = el.scrollHeight > el.clientHeight + 60;
          const likelyMain = r.width > 300 && r.height > 300;
          if (canScroll && likelyMain) {
            el.scrollTop = el.scrollHeight;
          }
        } catch(e) {}
      }
      return true;
    }
    """

    for i in range(1, 8):
        try:
            page.evaluate(scroll_js)
        except Exception as e:
            wlog(f"第 {i} 次滚动发布设置失败：{repr(e)}")

        step_wait(cfg, f"下滑发布页到底部 第{i}次")

        # 再用鼠标滚轮补充，确保是真实页面滚动
        try:
            page.mouse.wheel(0, 1200)
        except Exception:
            pass

        step_wait(cfg, f"鼠标滚轮下滑 第{i}次")

        if validate_publish_settings_visible(cfg, page, raise_error=False):
            wlog("已成功下滑到发布设置区域，底部发布按钮可见。")
            return True

    save_debug(cfg, page, "publish_settings_scroll_failed")
    raise RuntimeError("未能下滑到发布设置区域最底部。")


def get_text_rects(page, word):
    return page.evaluate("""
    (word) => {
      const walker=document.createTreeWalker(document.body, NodeFilter.SHOW_TEXT);
      const out=[];
      while(walker.nextNode()){
        const node=walker.currentNode, text=node.nodeValue||'', idx=text.indexOf(word);
        if(idx<0) continue;
        const range=document.createRange(); range.setStart(node,idx); range.setEnd(node,idx+word.length);
        for(const r of range.getClientRects()){
          if(r.width>0&&r.height>0&&r.top>0&&r.top<window.innerHeight){
            out.push({x:r.left,y:r.top,w:r.width,h:r.height,cx:r.left+r.width/2,cy:r.top+r.height/2});
          }
        }
      }
      out.sort((a,b)=>b.y-a.y);
      return out;
    }
    """, word) or []


def click_publish_mode(page, mode):
    word = "定时发布" if mode == "schedule" else "立即发布"
    alt = "定时" if mode == "schedule" else "直接发布"
    for attempt in range(1, 6):
        rects = get_text_rects(page, word) or get_text_rects(page, alt)
        if not rects:
            step_wait(reason=f"等待 {word} 出现")
            continue
        r = rects[0]
        points = [(r["x"]-18, r["cy"]), (r["x"]-30, r["cy"]), (r["cx"], r["cy"])]
        for x,y in points:
            if x < 0: continue
            page.mouse.click(float(x), float(y))
            step_wait(reason=f"点击 {word} 后等待")
            if mode == "schedule":
                if has_any(page, ["定时发布"], 500) and schedule_input_candidates(page):
                    return True
            else:
                return True
    return False


def schedule_input_candidates(page):
    try:
        return page.evaluate("""
        () => {
          function visible(el){ const r=el.getBoundingClientRect(); const s=getComputedStyle(el); return r.width>0&&r.height>0&&s.display!=='none'&&s.visibility!=='hidden'; }
          function txt(el){ return ((el.innerText||el.textContent||'')+'').trim(); }
          const labels=[...document.querySelectorAll('div,span,label,p')].filter(visible);
          const rects=[];
          for(const el of labels){ const t=txt(el); if(t.includes('发布时间')||t.includes('发布设置')){ const r=el.getBoundingClientRect(); rects.push({x:r.left,y:r.top,w:r.width,h:r.height}); } }
          const out=[];
          [...document.querySelectorAll('input')].filter(visible).forEach((el,idx)=>{
            const r=el.getBoundingClientRect(); const type=String(el.type||'').toLowerCase();
            if(['radio','checkbox','file','hidden','button','submit'].includes(type)) return;
            if(r.width<90||r.height<20) return;
            const val=el.value||'', ph=el.getAttribute('placeholder')||'', cls=String(el.className||'');
            let score=0;
            for(const pr of rects){
              const dy=Math.abs((r.top+r.height/2)-(pr.y+pr.h/2));
              if(dy<100 && r.top>pr.y-30) score+=60;
            }
            if(val.includes('202')) score+=35;
            if(ph.includes('日期')||ph.includes('时间')||ph.includes('发布')||cls.includes('picker')) score+=20;
            if(score>0) out.push({idx,x:r.left,y:r.top,w:r.width,h:r.height,value:val,score});
          });
          out.sort((a,b)=>b.score-a.score);
          return out.slice(0,6);
        }
        """) or []
    except Exception:
        return []


def set_schedule_time(page, slot):
    """
    V27：优化定时时间输入。
    日期选择保持原逻辑；时间不再滚动选择小时/分钟。
    根据实测：删除原有时分，再输入目标 HH:MM 即可正常。
    所以流程改为：
    1. 点击发布时间输入框打开面板。
    2. 点击目标日期。
    3. 对同一个时间输入框保留日期部分，删除原有 HH:MM，输入目标 HH:MM。
    4. 点击确认。
    """
    target = slot.strftime("%Y-%m-%d %H:%M")
    target_date = slot.strftime("%Y-%m-%d")
    target_time = slot.strftime("%H:%M")
    day_num = str(slot.day)

    cands = schedule_input_candidates(page)
    if not cands:
        wlog("未找到定时时间输入框。")
        return False

    cand = cands[0]

    # 点击发布时间输入框，打开日期时间面板
    try:
        page.mouse.click(float(cand["x"] + cand["w"] / 2), float(cand["y"] + cand["h"] / 2))
        step_wait(reason="点击发布时间输入框，打开日期面板")
    except Exception as e:
        wlog(f"点击时间输入框失败：{repr(e)}")
        return False

    # 点击日期，日期这一步当前已经稳定，保留
    try:
        res = page.evaluate("""
        ({day}) => {
          function visible(el){
            const r = el.getBoundingClientRect();
            const s = getComputedStyle(el);
            return r.width > 0 && r.height > 0 && s.display !== 'none' && s.visibility !== 'hidden';
          }
          function txt(el){ return ((el.innerText || el.textContent || '') + '').trim(); }

          const nodes = [...document.querySelectorAll('td,button,span,div')].filter(visible);
          const arr = [];
          for (const el of nodes) {
            const t = txt(el);
            if (t !== day) continue;
            const r = el.getBoundingClientRect();
            if (r.width > 90 || r.height > 90 || r.top < 160) continue;
            const cls = String(el.className || '').toLowerCase();
            let score = 0;
            if (cls.includes('disabled')) score -= 300;
            if (r.left > 250 && r.left < window.innerWidth * 0.8) score += 20;
            if (cls.includes('selected') || cls.includes('active') || cls.includes('today')) score += 5;
            arr.push({el, score, x:r.left, y:r.top});
          }
          arr.sort((a,b)=>b.score-a.score);
          if (arr.length) {
            arr[0].el.click();
            return {ok:true, count:arr.length};
          }
          return {ok:false, count:0};
        }
        """, {"day": day_num})
        wlog(f"点击日期结果：{res}")
    except Exception as e:
        wlog(f"点击日期失败：{repr(e)}")

    step_wait(reason="点击日期后等待")

    # 核心修复：保留日期，清掉旧时分，再输入目标 HH:MM
    # 不使用全页面 Ctrl+A；只操作最近的发布时间 input。
    try:
        result = page.evaluate("""
        ({x, y, targetDate, targetTime, target}) => {
          function visible(el){
            const r = el.getBoundingClientRect();
            const s = getComputedStyle(el);
            return r.width > 0 && r.height > 0 && s.display !== 'none' && s.visibility !== 'hidden';
          }

          const inputs = [...document.querySelectorAll('input')].filter(visible)
            .filter(el => !['radio','checkbox','file','hidden','button','submit'].includes(String(el.type||'').toLowerCase()))
            .filter(el => {
              const r = el.getBoundingClientRect();
              return r.width >= 90 && r.height >= 20;
            });

          let best = null;
          let bestD = 999999;
          for (const el of inputs) {
            const r = el.getBoundingClientRect();
            const cx = r.left + r.width / 2;
            const cy = r.top + r.height / 2;
            const d = Math.abs(cx - x) + Math.abs(cy - y);
            if (d < bestD) {
              bestD = d;
              best = el;
            }
          }

          if (!best) return {ok:false, reason:'input not found', value:''};

          best.scrollIntoView({block:'center'});
          best.focus();
          best.click();

          let current = best.value || '';

          // 先尽量保留已经选好的日期；如果输入框日期不对，直接用目标日期
          let datePart = targetDate;
          const m = current.match(/\\d{4}[-/]\\d{1,2}[-/]\\d{1,2}/);
          if (m) datePart = m[0].replace(/\\//g, '-');

          // 如果当前日期不是目标日期，仍以目标日期为准
          if (datePart !== targetDate) datePart = targetDate;

          const finalValue = datePart + ' ' + targetTime;

          const setter =
            Object.getOwnPropertyDescriptor(best.__proto__, 'value')?.set ||
            Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value')?.set;

          // 模拟"删除原时分，再输入新时分"：最终只改这个 input 的值，不动页面其他文字
          if (setter) setter.call(best, finalValue);
          else best.value = finalValue;

          best.dispatchEvent(new InputEvent('input', {bubbles:true, inputType:'insertText', data:finalValue}));
          best.dispatchEvent(new Event('change', {bubbles:true}));

          // 把光标放在最后，触发一次 Enter
          try {
            best.setSelectionRange(finalValue.length, finalValue.length);
          } catch(e) {}

          best.dispatchEvent(new KeyboardEvent('keydown', {bubbles:true, key:'Enter', code:'Enter'}));
          best.dispatchEvent(new KeyboardEvent('keyup', {bubbles:true, key:'Enter', code:'Enter'}));

          return {ok:true, oldValue:current, value:best.value || '', finalValue};
        }
        """, {
            "x": cand["x"] + cand["w"] / 2,
            "y": cand["y"] + cand["h"] / 2,
            "targetDate": target_date,
            "targetTime": target_time,
            "target": target,
        })
        wlog(f"删除原时分并输入目标时间结果：{result}")
    except Exception as e:
        wlog(f"输入目标时分失败：{repr(e)}")
        result = {"ok": False}

    step_wait(reason="输入目标时分后等待")

    click_text(page, ["确定", "确认", "完成"], timeout=2500)
    step_wait(reason="点击确认后等待")

    return verify_schedule_time(page, slot)


def verify_schedule_time(page, slot):
    target = slot.strftime("%Y-%m-%d %H:%M")
    vals = [str(x.get("value","")).strip() for x in schedule_input_candidates(page)]
    wlog(f"定时校验：目标={target}；输入框值={vals}")
    return any(v == target or v == target + ":00" or v.replace("/", "-") == target for v in vals)


def set_schedule(cfg, page, slot):
    scroll_to_publish_settings(cfg, page)
    if not click_publish_mode(page, "schedule"):
        save_debug(cfg, page, "schedule_mode_failed")
        raise RuntimeError("没有成功点击定时发布。")
    step_wait(cfg, "点击定时发布后等待")
    for _ in range(2):
        if set_schedule_time(page, slot):
            wlog("定时时间设置成功。")
            return
        step_wait(cfg, "定时时间未成功，重试")
    save_debug(cfg, page, "schedule_time_failed")
    raise RuntimeError("定时时间没有成功设置。")


def set_publish_now(cfg, page):
    """
    不定时发布逻辑：
    这里只是切换到"立即发布/不定时发布"的单选状态。
    最终提交仍然点击底部红色"发布"按钮，不存在单独的"立即发布"提交按钮。
    """
    scroll_to_publish_settings(cfg, page)
    if not click_publish_mode(page, "now"):
        save_debug(cfg, page, "publish_now_failed")
        raise RuntimeError("没有成功切换到立即发布/不定时发布状态。")
    step_wait(cfg, "切换立即发布状态后等待")



def publish_page_still_visible(page):
    """
    v1.1 修复：判断是否仍停留在发布编辑页。
    点击发布后不再点"确认/确定"，只观察页面是否跳转。
    """
    try:
        url = (page.url or "").lower()
    except Exception:
        url = ""
    if "content/post/image" in url or "creator-micro/content/upload" in url:
        return True

    try:
        body = page.locator("body").inner_text(timeout=1500)
    except Exception:
        body = ""

    publish_markers = ["基础信息", "作品描述", "发布设置", "选择音乐", "封面设置"]
    manage_markers = ["作品管理", "全部作品", "已发布", "审核中", "未通过"]

    if any(x in body for x in manage_markers) and not any(x in body for x in publish_markers[:2]):
        return False

    return any(x in body for x in publish_markers)


def manage_page_visible(page):
    """
    v1.1 修复：点击发布后，成功状态以进入作品管理页/作品列表为准。
    """
    try:
        url = (page.url or "").lower()
    except Exception:
        url = ""

    if "content/manage" in url or "creator-micro/content/manage" in url:
        return True

    try:
        body = page.locator("body").inner_text(timeout=1500)
    except Exception:
        body = ""

    if "作品管理" in body and ("全部作品" in body or "已发布" in body or "审核中" in body or "未通过" in body):
        return True

    # 定时作品发布后列表里常见"定时发布中"
    if "定时发布中" in body and ("作品管理" in body or "修改定时" in body):
        return True

    return False


def find_bottom_publish_button(page):
    """
    找底部真正的红色"发布"按钮。
    不找"立即发布"按钮，因为不定时发布也是点击底部"发布"。
    """
    try:
        buttons = page.locator("button").filter(has_text="发布")
        arr = []
        for i in range(min(buttons.count(), 30)):
            try:
                b = buttons.nth(i)
                if not b.is_visible(timeout=300):
                    continue
                box = b.bounding_box(timeout=500)
                if not box:
                    continue
                txt = (b.inner_text(timeout=500) or "").strip()
                if "发布" not in txt:
                    continue
                # 底部发布按钮通常在页面下方，优先取最后一个可见按钮
                arr.append((box.get("y", 0), b, txt))
            except Exception:
                pass
        if not arr:
            return None
        arr.sort(key=lambda x: x[0])
        return arr[-1][1]
    except Exception:
        return None


def click_bottom_publish_button(cfg, page):
    scroll_to_publish_settings(cfg, page)
    btn = find_bottom_publish_button(page)
    if btn is None:
        save_debug(cfg, page, "publish_button_not_found")
        raise RuntimeError("没有找到底部发布按钮。")
    btn.click(force=True, timeout=8000)
    step_wait(cfg, "点击发布后等待")
    return True


def wait_publish_result_once(cfg, page, wait_seconds=35):
    """
    点击发布后等待一次结果。
    返回：
    success：进入作品管理页或出现成功提示
    still_publish：仍停留在发布页，需要再次点击发布
    error：检测到失败提示
    unknown：未确定状态
    """
    end = time.time() + wait_seconds
    last_state_log = 0

    while time.time() < end:
        if manage_page_visible(page):
            wlog("发布结果判断：已进入作品管理页面，发布成功。")
            return "success"

        if has_any(page, ["发布成功", "定时发布成功", "提交成功", "作品已进入定时发布", "发布任务已提交"], 600):
            wlog("发布结果判断：检测到成功提示。")
            return "success"

        if has_any(page, ["发布失败", "错误", "请完善", "不能为空", "违规", "过于频繁", "稍后再试"], 600):
            save_debug(cfg, page, "publish_error")
            wlog("发布结果判断：检测到失败或错误提示。")
            return "error"

        if publish_page_still_visible(page):
            if time.time() - last_state_log > 8:
                wlog("发布结果判断：仍停留在发布页面，继续等待跳转。")
                last_state_log = time.time()
        else:
            if time.time() - last_state_log > 8:
                try:
                    wlog(f"发布结果判断：页面跳转中，当前URL={page.url}")
                except Exception:
                    wlog("发布结果判断：页面跳转中。")
                last_state_log = time.time()

        time.sleep(2)

    if manage_page_visible(page):
        return "success"

    if publish_page_still_visible(page):
        return "still_publish"

    return "unknown"


def submit(cfg, page):
    """
    v1.1.1：
    1. 点击发布后不再点击确认/确定。
    2. 如果仍停留在发布页面，会根据前端"失败重试次数"再次点击底部发布按钮。
    3. 不定时发布也是点击底部红色"发布"按钮，不寻找"立即发布"提交按钮。
    """
    retry_times = int(cfg.get("retry_times", 2) or 0)
    max_clicks = max(1, retry_times + 1)

    wlog(f"提交发布。最多点击发布 {max_clicks} 次。")

    for attempt in range(1, max_clicks + 1):
        wlog(f"第 {attempt}/{max_clicks} 次点击底部发布按钮。")
        click_bottom_publish_button(cfg, page)

        # 点击发布后不再点确认，只等待页面是否进入作品管理
        result = wait_publish_result_once(cfg, page, wait_seconds=35)

        if result == "success":
            return True

        if result == "error":
            return False

        if result == "still_publish":
            if attempt < max_clicks:
                wlog("发布后仍停留在发布页面，准备重试点击发布。")
                step_wait(cfg, "重试点击发布前等待")
                continue
            save_debug(cfg, page, "publish_still_on_publish_page")
            wlog("发布结果判断：达到前端设置的最大重试次数后，仍停留在发布页面。")
            return False

        # unknown：页面既不是发布页也不是管理页，短暂等待后再判断或重试
        if result == "unknown":
            if manage_page_visible(page):
                return True
            if publish_page_still_visible(page) and attempt < max_clicks:
                wlog("发布结果未确认且仍在发布页，准备重试点击发布。")
                continue
            save_debug(cfg, page, "publish_result_unknown")
            wlog("发布结果判断：未能确认发布是否成功。")
            return False

    return False



def keep_single_browser_tab(context, page=None):
    """
    V29：强制单标签模式。
    优先保留当前传入标签；如果没有，则保留最后一个创作者中心标签。
    关闭其它标签，避免越跑越多。
    """
    try:
        pages = list(context.pages)
        if not pages:
            return context.new_page()

        chosen = page if page in pages else None
        if chosen is None:
            douyin_pages = []
            for p in pages:
                try:
                    if "creator.douyin.com" in (p.url or ""):
                        douyin_pages.append(p)
                except Exception:
                    pass
            chosen = douyin_pages[-1] if douyin_pages else pages[-1]

        for p in list(pages):
            if p is chosen:
                continue
            try:
                p.close()
            except Exception:
                pass

        # v1.1：不主动拉起/恢复最小化浏览器
        safe_bring_to_front(ACTIVE_CFG or {}, chosen)
        wlog("已整理浏览器标签：仅保留 1 个发布标签。")
        return chosen
    except Exception as e:
        wlog(f"整理浏览器标签失败：{repr(e)}")
        try:
            return page or context.pages[-1]
        except Exception:
            return page




def wait_publish_interval(cfg, done, remaining_slots=None):
    """
    V1.1：每条发布成功后，进入下一条前等待一个可浮动发布间隔。
    这个等待和"每一步等待秒"分开，专门用于两条作品之间的间隔。
    """
    if remaining_slots is not None and remaining_slots <= 0:
        return
    try:
        a = float(cfg.get("publish_interval_min_seconds", 30) or 0)
        b = float(cfg.get("publish_interval_max_seconds", 90) or 0)
    except Exception:
        a, b = 30, 90

    if a < 0:
        a = 0
    if b < a:
        b = a

    if b <= 0:
        return

    d = random.uniform(a, b) if b > a else a
    wlog(f"发布间隔等待 {d:.1f} 秒后进入下一条。")
    start = time.time()
    while time.time() - start < d:
        wait_if_paused(cfg)
        time.sleep(min(1, d - (time.time() - start)))


def worker(config_path):
    global ACTIVE_CFG
    cfg = json.loads(Path(config_path).read_text(encoding="utf-8"))
    ACTIVE_CFG = cfg
    copies = read_copies(cfg)
    slots = build_slots(cfg)
    state = read_state(cfg)
    current_schedule_signature = schedule_signature(cfg)

    # v1.1 修复：如果日期/开始小时/结束小时/分钟点/每小时条数发生变化，
    # 说明用户修改了新排期，必须从第一个新排期开始，不能沿用旧 slot_index。
    if state.get("schedule_signature") != current_schedule_signature:
        old_si = int(state.get("slot_index", 0) or 0)
        state["slot_index"] = 0
        state["schedule_signature"] = current_schedule_signature
        write_state(cfg, state)
        wlog(f"检测到排期配置已变化，已重置定时序号：{old_si + 1} -> 1")

    ci = 0 if cfg.get("delete_copy_after_success", True) else int(state.get("copy_index", 0))
    si = int(state.get("slot_index", 0))
    max_run = int(cfg.get("max_posts_this_run", 0) or 0)
    done = 0
    retry_times = int(cfg.get("retry_times", 2) or 2)

    wlog(f"读取文案 {len(copies)} 条，当前文案序号：{ci+1}")
    wlog(f"排期 {len(slots)} 个，当前排期序号：{si+1}")

    open_browser_with_cdp(cfg, force_new=False)
    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp(f"http://127.0.0.1:{int(cfg['cdp_port'])}")
        context = browser.contexts[0] if browser.contexts else browser.new_context()
        page = context.pages[-1] if context.pages else context.new_page()
        try:
            page = keep_single_browser_tab(context, page)
            safe_bring_to_front(cfg, page)
        except Exception:
            pass

        while True:
            if max_run and done >= max_run:
                wlog(f"已达到本次运行数量限制：{max_run}")
                break
            imgs = list_images(cfg)
            if not imgs:
                alert_auto_pause("提醒：图片文件夹已没有可发布图片，脚本暂停。")
                append_log(cfg, {
                    "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "image": "",
                    "copy_index": ci + 1,
                    "schedule_time": "",
                    "copy_preview": "",
                    "status": "paused_no_images"
                })
                break

            copies = read_copies(cfg)
            if cfg.get("delete_copy_after_success", True):
                copies = filter_used_copies(copies, read_state(cfg))
            if not copies:
                alert_auto_pause("提醒：Excel 中已没有可发布文案，脚本暂停。")
                append_log(cfg, {
                    "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "image": str(imgs[0]) if imgs else "",
                    "copy_index": 0,
                    "schedule_time": "",
                    "copy_preview": "",
                    "status": "paused_no_copies"
                })
                break

            if cfg.get("delete_copy_after_success", True):
                ci = 0

            if ci >= len(copies):
                alert_auto_pause("提醒：文案序号已经超过当前文案数量，脚本暂停。可以重置发布进度或开启'发布成功后删除文案'。")
                append_log(cfg, {
                    "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "image": str(imgs[0]),
                    "copy_index": ci + 1,
                    "schedule_time": "",
                    "copy_preview": "",
                    "status": "paused_copy_index_out_of_range"
                })
                break

            if cfg.get("use_schedule", True) and si >= len(slots):
                alert_auto_pause(f"提醒：当前排期共 {len(slots)} 个，已执行到第 {si + 1} 个，说明当前填写的时间段已完成，脚本暂停。")
                append_log(cfg, {
                    "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "image": str(imgs[0]) if imgs else "",
                    "copy_index": ci + 1,
                    "schedule_time": "",
                    "copy_preview": "",
                    "status": "paused_schedule_slots_done"
                })
                break

            if cfg.get("random_image", False):
                img = random.choice(imgs)
            else:
                img = imgs[0]
            copy = copies[ci]
            slot = slots[si] if si < len(slots) else datetime.now()
            ok = False
            last_error = None

            for attempt in range(1, retry_times + 2):
                wlog("=" * 60)
                wlog(f"第 {attempt}/{retry_times+1} 次尝试：图片={img.name}；文案序号={ci+1}；定时={slot.strftime('%Y-%m-%d %H:%M') if cfg.get('use_schedule', True) else '立即发布'}")
                try:
                    page = keep_single_browser_tab(context, page)
                    enter_publish_page(cfg, page)
                    upload_image(cfg, page, img)
                    step_wait(cfg, "已选择图片，等待进入图文编辑页")
                    wait_after_upload_to_editor(cfg, page)

                    fill_copy(cfg, page, copy)
                    validate_copy_filled(cfg, page, copy)
                    # 文案已真实写入发布页后，立即标记为已使用，避免暂停/失败后重复使用。
                    mark_copy_used_in_state(cfg, copy)

                    if cfg.get("music_required", True):
                        select_music(cfg, page)
                        validate_music_added(cfg, page)
                    close_music_panel_if_open(cfg, page)

                    scroll_to_publish_settings(cfg, page)
                    validate_publish_settings_visible(cfg, page)

                    if cfg.get("use_schedule", True):
                        set_schedule(cfg, page, slot)
                        if not verify_schedule_time(page, slot):
                            raise RuntimeError("发布前定时时间校验失败。")
                    else:
                        set_publish_now(cfg, page)

                    # V21：发布前最终校验，图片是否上传完成只在这里判断
                    wait_uploaded_with_config(cfg, page)
                    validate_uploaded_image(cfg, page)
                    validate_copy_filled(cfg, page, copy)
                    if cfg.get("music_required", True):
                        validate_music_added(cfg, page)

                    ok = submit(cfg, page)
                    if not ok:
                        raise RuntimeError("提交发布后未确认成功。")
                    break
                except Exception as e:
                    last_error = e
                    wlog(f"第 {attempt} 次失败：{repr(e)}")
                    save_debug(cfg, page, f"attempt_{attempt}_failed")
                    if attempt <= retry_times:
                        try:
                            page.goto(cfg["creator_url"], wait_until="domcontentloaded", timeout=30000)
                        except Exception:
                            pass
                        step_wait(cfg, "准备重试当前作品")
                    else:
                        wlog("已达到最大重试次数，暂停。")

            if ok:
                if cfg.get("delete_image_after_success", True):
                    try:
                        img.unlink()
                        wlog(f"已删除已发布图片：{img.name}")
                    except Exception as e:
                        wlog(f"删除图片失败：{e}")

                if cfg.get("delete_copy_after_success", True):
                    try:
                        deleted_copy = delete_copy_from_excel(cfg, copy)
                        if not deleted_copy:
                            mark_copy_used_in_state(cfg, copy)
                    except Exception as e:
                        wlog(f"删除已用文案失败：{repr(e)}")
                        mark_copy_used_in_state(cfg, copy)

                append_log(cfg, {
                    "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "image": str(img),
                    "copy_index": ci + 1,
                    "schedule_time": slot.strftime("%Y-%m-%d %H:%M") if cfg.get("use_schedule", True) else "立即发布",
                    "copy_preview": copy[:80].replace("\n", " "),
                    "status": "success"
                })

                if cfg.get("delete_copy_after_success", True):
                    ci = 0
                else:
                    ci += 1

                si += 1
                done += 1
                state_now = read_state(cfg)
                write_state(cfg, {
                    "copy_index": ci,
                    "slot_index": si,
                    "used_copy_keys": state_now.get("used_copy_keys", []),
                    "schedule_signature": schedule_signature(cfg)
                })

                remaining_slots = None
                if cfg.get("use_schedule", True):
                    remaining_slots = len(slots) - si
                    if remaining_slots <= 0:
                        alert_auto_pause("提醒：当前填写的定时开始时间到结束时间内，所有发布时间已完成，脚本暂停。")
                    else:
                        wait_publish_interval(cfg, done, remaining_slots=remaining_slots)
                else:
                    wait_publish_interval(cfg, done)
            else:
                append_log(cfg, {
                    "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "image": str(img),
                    "copy_index": ci + 1,
                    "schedule_time": slot.strftime("%Y-%m-%d %H:%M") if cfg.get("use_schedule", True) else "立即发布",
                    "copy_preview": copy[:80].replace("\n", " "),
                    "status": "error_after_retries: " + repr(last_error)
                })
                break

    wlog("脚本运行结束。")


class App:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_NAME + "｜脚本启动器")
        self.root.geometry("1080x780")
        self.root.minsize(980, 720)
        try:
            if LOGO_ICO.exists():
                self.root.iconbitmap(str(LOGO_ICO))
        except Exception:
            pass
        try:
            self.root.configure(bg="#f4f6fb")
        except Exception:
            pass
        self.root.minsize(980, 720)
        try:
            if LOGO_ICO.exists():
                self.root.iconbitmap(str(LOGO_ICO))
        except Exception:
            pass
        try:
            self.root.configure(bg="#f4f6fb")
        except Exception:
            pass
        self.cfg = load_config()
        self.vars = {}
        self.bool_vars = {}
        self.proc = None
        self.q = queue.Queue()
        self.build()
        try:
            self.root.after(1500, self.auto_check_update)
        except Exception:
            pass

    def add_row(self, parent, row, label, key, browse=None, width=74):
        tk.Label(parent, text=label, width=18, anchor="e").grid(row=row, column=0, padx=5, pady=4, sticky="e")
        v = tk.StringVar(value=str(self.cfg.get(key, "")))
        self.vars[key] = v
        tk.Entry(parent, textvariable=v, width=width).grid(row=row, column=1, padx=5, pady=4, sticky="we")
        if browse == "file":
            tk.Button(parent, text="选择文件", command=lambda: self.choose_file(key)).grid(row=row, column=2, padx=5)
        elif browse == "folder":
            tk.Button(parent, text="选择文件夹", command=lambda: self.choose_folder(key)).grid(row=row, column=2, padx=5)
        else:
            tk.Label(parent, text="").grid(row=row, column=2)
        return row + 1

    def build(self):
        top = tk.LabelFrame(self.root, text="基础配置")
        top.pack(fill="x", padx=10, pady=6)
        top.columnconfigure(1, weight=1)
        r = 0
        r = self.add_row(top, r, "浏览器快捷方式/EXE", "browser_path", "file")
        r = self.add_row(top, r, "图片文件夹", "image_dir", "folder")
        r = self.add_row(top, r, "文案Excel", "excel_path", "file")
        r = self.add_row(top, r, "工作表名", "sheet_name")
        r = self.add_row(top, r, "浏览器用户目录(可空)", "browser_user_data_dir", "folder")
        r = self.add_row(top, r, "浏览器Profile目录(可空)", "browser_profile_directory")

        mid = tk.LabelFrame(self.root, text="定时/等待配置")
        mid.pack(fill="x", padx=10, pady=4)
        fields = [
            ("开始日期", "publish_start_date"), ("结束日期", "publish_end_date"), ("开始小时", "start_hour"), ("结束小时0-24", "end_hour"), ("每小时条数", "per_hour_count"),
            ("分钟点", "custom_minutes"), ("每步最小等待秒", "wait_min_seconds"), ("每步最大等待秒", "wait_max_seconds"), ("失败重试次数", "retry_times"),
            ("发布间隔最小秒", "publish_interval_min_seconds"), ("发布间隔最大秒", "publish_interval_max_seconds"),
            ("上传检测间隔秒", "upload_check_interval_seconds"), ("上传最大等待秒", "upload_max_wait_seconds"),
            ("话题识别等待秒", "topic_wait_seconds"), ("创作者中心等待秒", "creator_center_wait_seconds"), ("调试端口", "cdp_port"), ("本次最多发几条 0不限", "max_posts_this_run"),
        ]
        for i, (lab, key) in enumerate(fields):
            tk.Label(mid, text=lab).grid(row=i//4, column=(i%4)*2, padx=5, pady=5, sticky="e")
            v = tk.StringVar(value=str(self.cfg.get(key, "")))
            self.vars[key] = v
            tk.Entry(mid, textvariable=v, width=15).grid(row=i//4, column=(i%4)*2+1, padx=5, pady=5, sticky="w")

        opt = tk.LabelFrame(self.root, text="运行选项")
        opt.pack(fill="x", padx=10, pady=4)
        opts = [
            ("发布成功后删除图片", "delete_image_after_success"),
            ("发布成功后删除文案", "delete_copy_after_success"),
            ("使用定时发布", "use_schedule"),
            ("随机使用图片", "random_image"),
            ("复用已存在调试端口", "reuse_existing_cdp"),
            ("启动前关闭Chrome残留", "close_chrome_before_start"),
            ("后台运行不拉起浏览器", "no_raise_browser"),
        ]
        for i, (text, key) in enumerate(opts):
            bv = tk.BooleanVar(value=bool(self.cfg.get(key, True)))
            self.bool_vars[key] = bv
            tk.Checkbutton(opt, text=text, variable=bv).grid(row=i//3, column=i%3, padx=8, pady=5, sticky="w")

        paths = tk.LabelFrame(self.root, text="状态/日志路径")
        paths.pack(fill="x", padx=10, pady=4)
        paths.columnconfigure(1, weight=1)
        r = 0
        r = self.add_row(paths, r, "进度文件", "state_path", "file")
        r = self.add_row(paths, r, "发布日志CSV", "log_path", "file")
        r = self.add_row(paths, r, "调试截图文件夹", "debug_dir", "folder")

        buttons = tk.Frame(self.root)
        buttons.pack(fill="x", padx=10, pady=6)
        tk.Button(buttons, text="保存配置", command=self.save_from_ui, width=12).pack(side="left", padx=4)
        tk.Button(buttons, text="打开/连接浏览器", command=self.open_browser_window, width=16, bg="#14a44d", fg="white").pack(side="left", padx=4)
        tk.Button(buttons, text="开始发布任务", command=self.start, width=16, bg="#2d8cff", fg="white").pack(side="left", padx=4)
        tk.Button(buttons, text="暂停/继续", command=self.toggle_pause, width=12).pack(side="left", padx=4)
        tk.Button(buttons, text="停止任务", command=self.stop, width=12).pack(side="left", padx=4)
        tk.Button(buttons, text="清空窗口日志", command=lambda: self.logbox.delete("1.0", "end"), width=12).pack(side="left", padx=4)
        tk.Button(buttons, text="检查更新", command=self.check_update_click, width=12).pack(side="left", padx=4)
        tk.Button(buttons, text="打开程序文件夹", command=lambda: os.startfile(str(APP_DIR)), width=14).pack(side="left", padx=4)
        tk.Button(buttons, text="重置发布进度", command=self.reset_state, width=14).pack(side="left", padx=4)

        self.logbox = ScrolledText(self.root, height=22, font=("Consolas", 10))
        self.logbox.pack(fill="both", expand=True, padx=10, pady=6)
        self.write_ui("抖音单图文发布v2.2.1 已启动。已增加在线更新：启动自动检查，也可点击“检查更新”。\n")


    def auto_check_update(self):
        self.check_update_async(auto=True)

    def check_update_click(self):
        self.check_update_async(auto=False)

    def check_update_async(self, auto=False):
        def task():
            try:
                info = available_update_info()
                self.root.after(0, lambda: self.handle_update_result(info, auto))
            except Exception as e:
                if not auto:
                    self.root.after(0, lambda: messagebox.showerror("检查更新失败", str(e)))
        threading.Thread(target=task, daemon=True).start()

    def handle_update_result(self, info, auto=False):
        if not info:
            if not auto:
                messagebox.showinfo("检查更新", f"当前已经是最新版本：v{APP_VERSION}")
            return
        latest = str(info.get("latest_version") or info.get("version") or "")
        notes = format_release_notes(info)
        text = f"发现新版本：v{APP_VERSION} → v{latest}\n\n"
        if notes:
            text += notes + "\n\n"
        text += "是否现在下载并更新？更新时会保留本机配置 app/douyin_gui_config.json。"
        if messagebox.askyesno("发现新版本", text):
            self.start_online_update(info)

    def start_online_update(self, info):
        if self.proc and self.proc.poll() is None:
            messagebox.showwarning("正在运行", "发布任务正在运行，请先停止任务后再更新。")
            return
        try:
            self.save_from_ui()
        except Exception:
            pass
        updater = APP_DIR / "online_updater.py"
        if not updater.exists():
            messagebox.showerror("无法更新", f"缺少更新器文件：{updater}")
            return
        preserve = info.get("preserve_paths") or PRESERVE_UPDATE_PATHS
        if isinstance(preserve, list):
            preserve_arg = json.dumps(preserve, ensure_ascii=False)
        else:
            preserve_arg = json.dumps(PRESERVE_UPDATE_PATHS, ensure_ascii=False)
        cmd = [
            worker_python_executable(),
            str(updater),
            "--install-dir", str(APP_DIR.parent),
            "--current-pid", str(os.getpid()),
            "--download-url", str(info.get("download_url", "")),
            "--sha256", str(info.get("sha256", "")),
            "--launch", str(Path(__file__).resolve()),
            "--preserve", preserve_arg,
        ]
        try:
            subprocess.Popen(cmd, cwd=str(APP_DIR), creationflags=no_window_creationflags())
            messagebox.showinfo("开始更新", "更新器已启动。当前窗口会关闭，更新完成后会自动重新打开。")
            self.root.after(300, self.root.destroy)
        except Exception as e:
            messagebox.showerror("启动更新失败", repr(e))

    def choose_file(self, key):
        if key == "excel_path":
            p = filedialog.askopenfilename(
                title="选择文案Excel",
                filetypes=[("Excel 文案表", "*.xlsx *.xls *.xlsm"), ("所有文件", "*.*")],
            )
        else:
            p = filedialog.askopenfilename()
        if p:
            self.vars[key].set(p)

    def choose_folder(self, key):
        p = filedialog.askdirectory()
        if p:
            self.vars[key].set(p)

    def collect_cfg(self):
        cfg = self.cfg.copy()
        int_keys = {"start_hour","end_hour","per_hour_count","cdp_port","max_posts_this_run","retry_times"}
        float_keys = {"wait_min_seconds","wait_max_seconds","publish_interval_min_seconds","publish_interval_max_seconds","upload_check_interval_seconds","upload_max_wait_seconds","topic_wait_seconds"}
        for k, v in self.vars.items():
            val = v.get().strip()
            if k in int_keys:
                try: cfg[k] = int(val)
                except Exception: cfg[k] = 0
            elif k in float_keys:
                try: cfg[k] = float(val)
                except Exception: cfg[k] = 2.0
            else:
                cfg[k] = val
        for k, bv in self.bool_vars.items():
            cfg[k] = bool(bv.get())

        # V27：这些是脚本运行必需项，不再显示到前端让用户勾选，后台固定启用
        cfg["music_required"] = True
        cfg["keep_browser_open"] = True
        cfg["auto_switch_graphic_mode"] = True
        cfg["platform_music_only"] = True
        return cfg

    def save_from_ui(self):
        self.cfg = self.collect_cfg()
        save_config(self.cfg)
        self.write_ui("配置已保存。下一次点击'开始发布任务'会立刻使用当前最新配置。\n")

    def write_ui(self, text):
        self.logbox.insert("end", text)
        self.logbox.see("end")

    def reader(self):
        try:
            for line in self.proc.stdout:
                self.q.put(line)
        except Exception as e:
            self.q.put("读取日志失败：" + repr(e) + "\n")

    def poll(self):
        while not self.q.empty():
            self.write_ui(self.q.get())
        if self.proc and self.proc.poll() is None:
            self.root.after(200, self.poll)
        elif self.proc:
            self.write_ui(f"\n进程已结束，退出码：{self.proc.returncode}\n")
            self.proc = None

    def run_subprocess(self, args):
        env = os.environ.copy()
        env["PYTHONIOENCODING"] = "utf-8"
        env["PYTHONUTF8"] = "1"
        self.proc = subprocess.Popen(args, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                                     text=True, encoding="utf-8", errors="replace",
                                     bufsize=1, cwd=str(APP_DIR), env=env,
                                     creationflags=no_window_creationflags())
        threading.Thread(target=self.reader, daemon=True).start()
        self.root.after(200, self.poll)

    def open_browser_window(self):
        self.save_from_ui()
        if self.proc and self.proc.poll() is None:
            messagebox.showwarning("正在运行", "脚本已经在运行。")
            return
        cmd = [worker_python_executable(), str(Path(__file__).resolve()), "--open-browser", str(CONFIG_PATH)]
        self.write_ui("\n打开浏览器窗口命令：" + " ".join(cmd) + "\n")
        self.run_subprocess(cmd)

    def start(self):
        # V1.1：每次开始发布前，都强制保存并重新加载当前前端配置。
        # 这样停止任务后修改配置并保存，再开始时一定运行新配置。
        if self.proc and self.proc.poll() is None:
            messagebox.showwarning("正在运行", "脚本已经在运行。如需修改配置，请先点击停止任务。")
            return

        if self.proc and self.proc.poll() is not None:
            self.proc = None

        self.save_from_ui()
        self.cfg = load_config()

        try:
            flag = pause_flag_path(self.cfg)
            if flag.exists():
                flag.unlink()
        except Exception:
            pass

        cmd = [worker_python_executable(), str(Path(__file__).resolve()), "--worker", str(CONFIG_PATH)]
        self.write_ui("启动命令：" + " ".join(cmd) + "\n")
        self.write_ui("已加载最新保存配置并开始发布任务。\n")

        self.proc = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            creationflags=no_window_creationflags()
        )
        threading.Thread(target=self.reader, daemon=True).start()
        self.root.after(200, self.poll)

    def toggle_pause(self):
        self.save_from_ui()
        flag = pause_flag_path(self.cfg)
        try:
            flag.parent.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass

        if flag.exists():
            try:
                flag.unlink()
            except Exception:
                pass
            self.write_ui("已继续脚本。\n")
        else:
            try:
                flag.write_text("paused", encoding="utf-8")
            except Exception:
                pass
            self.write_ui("已暂停脚本。当前步骤完成后会停在下一步前等待。\n")

    def stop(self):
        try:
            if self.proc and self.proc.poll() is None:
                self.proc.terminate()
                try:
                    self.proc.wait(timeout=5)
                except Exception:
                    try:
                        self.proc.kill()
                        self.proc.wait(timeout=3)
                    except Exception:
                        pass
                self.write_ui("已发送停止命令，当前任务已停止。\n")
            else:
                self.write_ui("当前没有正在运行的任务。\n")
        except Exception as e:
            self.write_ui("停止任务失败：" + repr(e) + "\n")
        finally:
            self.proc = None

    def reset_state(self):
        self.save_from_ui()
        p = Path(self.cfg["state_path"])
        if messagebox.askyesno("确认", f"确定重置发布进度吗？\n{p}"):
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_text(json.dumps({"copy_index":0,"slot_index":0,"used_copy_keys":[],"schedule_signature":""}, ensure_ascii=False, indent=2), encoding="utf-8")
            self.write_ui("发布进度已重置。\n")


def run_gui():
    if tk is None:
        print("当前 Python 不支持 tkinter，请安装标准版 Python。")
        return
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    if "--worker" in sys.argv:
        worker(sys.argv[sys.argv.index("--worker") + 1])
    elif "--open-browser" in sys.argv:
        cfg = json.loads(Path(sys.argv[sys.argv.index("--open-browser") + 1]).read_text(encoding="utf-8"))
        open_browser_with_cdp(cfg, force_new=False)
        wlog("浏览器前端窗口已打开。")
    else:
        run_gui()
